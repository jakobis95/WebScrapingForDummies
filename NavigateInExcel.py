from openpyxl import Workbook
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from openpyxl.formatting.rule import ColorScale, FormatObject, CellIsRule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle

def conditional_formatting_with_rules(ws, todayCol):
    #styles
    greyFill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    greenFill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    redFill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    whiteFont = Font(color='FFFFFF')
    greenFont = Font(color='006100')
    redFont = Font(color='9C0006')


    first = FormatObject(type='num', val=50)
    last = FormatObject(type='num', val=65)
    # colors match the format objects:
    colors = [Color('00B0F0'), Color('FF0000')]
    cs2 = ColorScale(cfvo=[first, last], color=colors)
    # create a rule with the color scale
    from openpyxl.formatting.rule import Rule
    rule = Rule(type='colorScale', colorScale=cs2)
    FirstCellIndex = ws.cell(row=2, column=todayCol).coordinate
    LastCellIndex = ws.cell(row=ws.max_row, column=todayCol).coordinate
    print(str(FirstCellIndex) + ':' + str(LastCellIndex))
    area = str(FirstCellIndex) + ':' + str(LastCellIndex)
    ws.conditional_formatting.add(area,CellIsRule(operator='greaterThan', formula=['101'], fill=greyFill, font=whiteFont))
    ws.conditional_formatting.add(area, rule)


    #Formatierung differenz zum Vortag
    FirstCellIndex = ws.cell(row=2, column=todayCol-1).coordinate
    LastCellIndex = ws.cell(row=ws.max_row, column=todayCol-1).coordinate
    print(str(FirstCellIndex) + ':' + str(LastCellIndex))
    divarea = str(FirstCellIndex) + ':' + str(LastCellIndex)
    ws.conditional_formatting.add(divarea, CellIsRule(operator='between', formula=['1', '99'], fill=redFill, font=redFont))
    ws.conditional_formatting.add(divarea, CellIsRule(operator='between', formula=['-1', '-99'], fill=greenFill, font=greenFont))
    ws.conditional_formatting.add(divarea, CellIsRule(operator='greaterThan', formula=['100'], fill=greyFill, font=whiteFont))
    ws.conditional_formatting.add(divarea, CellIsRule(operator='lessThan', formula=['-100'], fill=greyFill, font=whiteFont))

def searchXL(ws, searchTerm, searchArea = 0, rowcol = "all", begin = 0): # ws=Worksheet, searchTerm, searchArea= specifies the row or column that should be searched, rowcol= if you want to search a row or a column,begin is used to describe from where to start searching in a row or column
    Ycor = "notFound"
    Xcor = "notFound"

    #search area parameters
    if searchArea != 0 or rowcol != "all":
        #search only one row
        if rowcol == "row":
            minRow = searchArea
            maxRow = minRow
            minCol = begin
            maxCol = ws.max_column
        #search only one col
        elif rowcol =="col":
            minCol = searchArea
            maxCol = minCol
            minRow = begin
            maxRow = ws.max_row
    #search whole Worksheet
    else:
        minCol = 0
        minRow = 0
        maxCol = ws.max_column
        maxRow = ws.max_row


    #print("Fehlerhaft am " + datetime.today().strftime('%d.%m.'))
    # for row in ws.iter_rows(max_col=1, max_row=10):
    #     for cell in row:
    #         print(cell.value)
    #         if cell.value == "ID":
    #             print(cell.column)  # change column number for any cell value you want
    #             print("break")
    #             break

    # Interrates over search area
    for col in ws.iter_cols(min_row=minRow, max_row=maxRow, min_col=minCol, max_col= maxCol):
        for cell in col:
            if cell.value == searchTerm:
                Ycor = cell.row
                Xcor = cell.column
                #print("Suchbegriff an Coordinate[" + str(cell.row) + "," + str(cell.column) + "] gefunden" )  # change column number for any cell value you want
                break

    #if Xcor == "notFound":
        #print("nichts gefunden")

    return Ycor, Xcor



if __name__ == "__main__":
    file = "PythonZuExcel.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    TodayCol = ws.max_column
    conditional_formatting_with_rules(ws, TodayCol)
    wb.save('PythonZuExcel.xlsx')
