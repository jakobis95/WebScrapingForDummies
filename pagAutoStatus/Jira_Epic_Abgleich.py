import json
import os
from openpyxl import load_workbook, styles
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL
from datetime import datetime
import pandas as pd
def update_HVAC_Status(jira_Worksheet, destination_Worksheet, index_column, destination_update_column, destination_index_column, type):
    i = 1
    coordinates = searchXL(jira_Worksheet, index_column, begin=1, rowcol="row", krit=True)
    index_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Summary", begin=1, rowcol="row")
    summary_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Issue Type", begin=1, rowcol="row")
    issue_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Epic Link", begin=1, rowcol="row", krit=True)
    epic_column = coordinates[1]
    coordinates = searchXL(destination_Worksheet, destination_update_column)
    destination_update_column = coordinates[1]
    coordinates = searchXL(destination_Worksheet, destination_index_column)
    destination_index_column = coordinates[1]
    print(type, "###############################")

    i = 2
    missing_Epics = []
    while jira_Worksheet.cell(row=i, column=index_column).value != None:
        index = jira_Worksheet.cell(row=i, column=index_column).value
        coordinates = searchXL(destination_Worksheet, index, begin=destination_index_column, rowcol="col")
        index_row = coordinates[0]
        if index_row != "notFound":
            print("Yes")
            #destination_Worksheet.cell(row=index_row, column=destination_update_column).value = "Ja"
        else:
            string = "Index", jira_Worksheet.cell(row=i, column=index_column).value, "Standort: ", jira_Worksheet.cell(row=i, column=summary_column).value
            missing_Epics.append(string)
        i = i + 1
    for epic in missing_Epics:
        print(epic)


if __name__ == "__main__":
#Todo Alle leeren Felder mit Nein befüllen
    UserName = os.getlogin()
    jira_xlsx_path_HVAC = "C:\\Users\\" + str(UserName) + "\\Downloads\\JiraEpics.xlsx"
    master_xlsx_path = "C:\\Users\\" + str(UserName) + "\\Downloads\\Aktuelle_IBN_Kopie.xlsx"
#Worksheets
    wb_master = load_workbook(filename=master_xlsx_path)
    status_ws = wb_master["STATUS"]
    wb_jira = load_workbook(filename=jira_xlsx_path_HVAC)
    jira_ws_HVAC = wb_jira["Sheet1"]
#Skriptaufrufe
    update_HVAC_Status(jira_ws_HVAC, status_ws, "Key", "HVAC Maßnahme\numgesetzt (WMQ1)", "Epic Ticket Nummer", "HVAC")
    os.startfile(master_xlsx_path)