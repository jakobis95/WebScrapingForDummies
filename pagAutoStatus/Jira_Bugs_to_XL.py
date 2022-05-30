import json
import os
from openpyxl import load_workbook, styles
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL
from datetime import datetime
import pandas as pd
def create_active_issue_String(JiraBugs, i, columnDict):
    Date = str(JiraBugs.cell(row=i, column=columnDict["Updated"]).value[:5])
    IssueKey = str(JiraBugs.cell(row=i, column=columnDict["Key"]).value)
    Summary = JiraBugs.cell(row=i, column=columnDict["Summary"]).value
    while Summary.find("]") >= 0:
        Index = Summary.find("]")
        Summary = Summary[Index + 1:]
    Summary = str(Summary)
    Status = str(JiraBugs.cell(row=i, column=columnDict["Status"]).value)
    Stringg = Date + ": " + IssueKey + ": "  + Summary +": Status = " + Status
    #print(Stringg)
    return Stringg

def write_Bugs_to_XL(jira_xlsxPfad, xlsxPfad):
    wb = load_workbook(filename=xlsxPfad)
    Status = wb["StatusKurz"]
    wbMaster = load_workbook(filename=jira_xlsxPfad)
    JiraBugs = wbMaster["Sheet1"]
    activeTicketsCord = searchXL(Status, "aktive Tickets und Errors")
    activeTicketsColumn = activeTicketsCord[1]
    activeTicketsRow = activeTicketsCord[0]
    wbEpicNumber = searchXL(Status,"Epic Ticket Nummer", activeTicketsCord[0],"row")
    wbEpicNumberCol = wbEpicNumber[1]
    print("activeTicket Col", activeTicketsColumn)
    EpicLink = searchXL(JiraBugs, "Epic Link")
    EpicLinkColumn = EpicLink[1]
    columnDict ={}
    c = 1
#Erase Active Bugs Column
    while Status.cell(row=activeTicketsRow + c, column=5).value != None:
        Status.cell(row=activeTicketsRow + c, column=activeTicketsColumn).value = None
        c= c +1
    print("active Issue Columne clean and ready")
    i=1
    while JiraBugs.cell(row=EpicLink[0], column=i).value != None:
        print(JiraBugs.cell(row=1, column=i).value)
        columnDict[JiraBugs.cell(row=1, column=i).value] = i
        i = i + 1
    print(columnDict)
    i = 1
    while JiraBugs.cell(row=i, column= columnDict["Epic Link"]).value != None:
        searchResponse = searchXL(Status, JiraBugs.cell(row=i, column=columnDict["Epic Link"]).value, wbEpicNumberCol, "col")
        if searchResponse[0] != "notFound":
            #print(JiraBugs.cell(row=i, column=columnDict["Epic Link"]).value, "found in Row:", searchResponse[0])
            if Status.cell(row=searchResponse[0] ,column=activeTicketsColumn ).value != None:
                FirstValue = Status.cell(row=searchResponse[0] ,column=activeTicketsColumn ).value
                SummaryStr = str(FirstValue) + create_active_issue_String(JiraBugs, i, columnDict) + "\n"
            else:
                SummaryStr = create_active_issue_String(JiraBugs, i, columnDict) + "\n"
            #SummaryStr = "\n" + str(JiraBugs.cell(row=i, column=columnDict["Summary"]).value)
            Status.cell(row=searchResponse[0] ,column=activeTicketsColumn ).value = SummaryStr
            Status.cell(row=searchResponse[0] ,column=activeTicketsColumn ).alignment = Alignment(wrapText=True, vertical='top')
        else:
            print("Not found ",JiraBugs.cell(row=i, column=columnDict["Epic Link"]).value, JiraBugs.cell(row=i, column=columnDict["Summary"]).value, create_active_issue_String(JiraBugs, i, columnDict))
        i = i + 1

    wb.save(xlsxPfad)
if __name__ == "__main__":
    UserName = os.getlogin()
    jira_CSV_Path = "C:\\Users\\" + str(UserName) + "\\Downloads\\CurrentBugs18052022.xlsx"
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Downloads\\IBN_SANDbox_Complete.xlsx"
    write_Bugs_to_XL(jira_CSV_Path, xlsxPfad)
