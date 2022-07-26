import json
import os
import msoffcrypto
import io
from openpyxl import load_workbook, styles
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL
from datetime import datetime
import pandas as pd
def update_HVAC_Status(jira_Worksheet, destination_Worksheet, index_column, destination_update_column, destination_index_column, type):
    i = 1

    coordinates = searchXL(jira_Worksheet, index_column, begin=1, rowcol="row", krit=True)
    index_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Summary", begin=1, rowcol="row")
    summary_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Issue Type", begin=1, rowcol="row")
    issue_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Epic Link", begin=1, rowcol="row", krit=True)
    epic_column = coordinates[1]
    coordinates = searchXL(destination_Worksheet, destination_update_column)
    destination_update_column = coordinates[1]
    coordinates = searchXL(destination_Worksheet, destination_index_column)
    destination_index_column = coordinates[1]
    print(type, "###############################")
    while jira_Worksheet.cell(row=i, column=issue_column).value != None:
        if jira_Worksheet.cell(row=i, column=issue_column).value != "Epic":
            jira_Worksheet.cell(row=i, column=index_column).value = jira_Worksheet.cell(row=i, column=epic_column).value
        i = i + 1

    i = 2
    while jira_Worksheet.cell(row=i, column=index_column).value != None:
        index = jira_Worksheet.cell(row=i, column=index_column).value
        coordinates = searchXL(destination_Worksheet, index, begin=destination_index_column, rowcol="col")
        index_row = coordinates[0]
        if index_row != "notFound":
            destination_Worksheet.cell(row=index_row, column=destination_update_column).value = "yes"
        else:
            print("Index not found", jira_Worksheet.cell(row=i, column=index_column).value, "Standort: ", jira_Worksheet.cell(row=i, column=summary_column).value)
        i = i + 1



if __name__ == "__main__":
#Todo Alle leeren Felder mit Nein befüllen
    UserName = os.getlogin()
    jira_xlsx_path_HVAC = "C:\\Users\\" + str(UserName) + "\\Downloads\\HVACOverview.xlsx"
    jira_xlsx_path_VR16 = "C:\\Users\\" + str(UserName) + "\\Downloads\\VR16UpdatedStations.xlsx"
    master_xlsx_path = "C:\\Users\\" + str(UserName) + "\\Downloads\\IBN_Tracking.xlsx"
#Open encrypted File
# decrypted_workbook = io.BytesIO()
#
# with open(master_xlsx_path, 'rb') as file:
#     office_file = msoffcrypto.OfficeFile(file)
#     office_file.load_key(password='Lademeister')
#     office_file.decrypt(decrypted_workbook)

# `filename` can also be a file-like object.
    #wb_master = load_workbook(filename=decrypted_workbook)
#Worksheets
    wb_master = load_workbook(filename=master_xlsx_path)
    status_ws = wb_master["STATUS"]
    wb_jira = load_workbook(filename=jira_xlsx_path_HVAC)
    jira_ws_HVAC = wb_jira["Sheet1"]
    wb_jira = load_workbook(filename=jira_xlsx_path_VR16)
    jira_ws_VR16 = wb_jira["Sheet1"]
#Skriptaufrufe
    update_HVAC_Status(jira_ws_HVAC, status_ws, "Key", "HVAC action (WMQ1)", "Jira Epic", "HVAC")
    update_HVAC_Status(jira_ws_VR16, status_ws, "Key", "VR 16 Update (WMQ2)", "Jira Epic", "VR16")
    wb_master.save(master_xlsx_path)
    os.startfile(master_xlsx_path)