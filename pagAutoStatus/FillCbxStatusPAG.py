import json
import os
from openpyxl import load_workbook, styles
from openpyxl.styles import Font, Color
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL
from datetime import datetime
from UserInterface.AddC import addColToRef,addCol, update, replaceTab

def updateChangeFormula(coordinateM1, coordinateToday):
    return '=IF(MID('+ str(coordinateToday) +',1,2)=MID('+ str(coordinateM1) +',1,2),"",IF(AND('+ str(coordinateToday) +'="yes",'+ str(coordinateM1) +'<>"yes"),"better",IF(AND('+ str(coordinateToday) +'<>"",'+ str(coordinateToday) +'<>"yes"),"worse","")))'

def addTodayColumn(wb, StatusWB, TodayColumn, TodayRow,CW,TableName):
    addCol(StatusWB, TodayColumn + 1, TodayRow, "Full/Part KW" + str(CW[1] + 1))
    addColToRef(StatusWB.tables[TableName].ref, 1)
    newRef = addColToRef(StatusWB.tables[TableName].ref, 1)
    replaceTab(StatusWB, newRef, TableName)
    wb.save(xlsxPfad)
    wb, StatusWB = update(xlsxPfad, "STATUS")
    return wb, StatusWB

def pOrN(StatusLastWeek, StatusThisWeek):
    if StatusThisWeek == None:
        StatusThisWeek = "no / offline"
    if StatusLastWeek == None:
        StatusLastWeek = "no / offline"
    Zustand = {"yes": 1, "no / part": 2, "no / full": 3, "no / offline" : 4 }
    change = Zustand[StatusLastWeek] - Zustand[StatusThisWeek]
    if change >= 0:
        if change < 1:
            return "Gleich"
        else:
            return "Besser"
    else:
        return "Schlechter"

def createCPlist(xlsxPfad, fehlerCBX):
    cpidLookup = list()
    for item in fehlerCBX:
        CPID = item['uniqueId']
        CPID = CPID[:12]
        print(CPID)
        cpidLookup.append(CPID)
    return cpidLookup

def defineFullOrPart(cpidLookup,cpid,cpNo):
    cpid = cpid[:12]
    cpsAmount = cpidLookup.count(cpid)
    if cpNo == None:
        cpNo = 2
    if cpsAmount < cpNo:
        Status = "no / part"
        return Status
    else:
        Status = "no / full"
        return Status


def WriteStatusToXL(xlsxPfad, offlineCBX, fehlerCBX):
    ###Styles
    my_yellow = styles.colors.Color(rgb='ffff00')
    my_yellow_letters = Font(color="FF0000")
    my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
    no_fill = styles.PatternFill(fill_type=None)
    ###
    cpidLookup = createCPlist(xlsxPfad, fehlerCBX)
    cp = 0
    wb = load_workbook(filename=xlsxPfad)
    todayCbxCounter = 1
    StatusWB = wb["STATUS"]
    TodayFehlerWB = wb["overviewToday"]
    #NIBfehlerWB = wb["NIBfehlerhaft"] #NIB steht für "nicht im backend"
    #NIBofflineWB = wb["NIBoffline"]
    commission = searchXL(StatusWB, 'Commissioning\n finalized')
    commissionColumn = commission[1]
    location = searchXL(StatusWB, "Location")
    locationCol = location[1]
    change = searchXL(StatusWB, 'Veränderung\nVorwoche') #todo sprache ändern auch in der XL
    changeCol = change[1]
    LastRow = StatusWB.max_row
    CW = datetime.today().isocalendar()
    TodayColumnString = "Full/Part KW" + str(CW[1]) #todo gibt diese den Namen der Woche an.
    #TodayColumnString = "Full/Part KW"
    #finds Column with the date of today
    cpNoCoordinate = searchXL(StatusWB, "hardware_prim_dc_no_chargers")
    cpNoColumn = cpNoCoordinate[1]
    CPMID = searchXL(StatusWB, "CPM ID")
    cpmidColumn = CPMID[1]

    TodayRow, TodayColumn = searchXL(StatusWB, TodayColumnString) #findet jetzt die heutige Spalte

    while TodayRow == "NotFound":
        input("Für Woche" + str(CW[1]) + " gibt es keine Spalte in der Exceltabelle\n Bitte Spalte hinzufügen, Excel Speichern und wieder schließen\n Bitte mit y bestätigen")
        wb, StatusWB = update(xlsxPfad, "STATUS")
        TodayRow, TodayColumn = searchXL(StatusWB, TodayColumnString)

    if TodayColumn != "notFound":
        print("XX today Column is:", TodayColumn)
####################################################################Fill in all destinations that are currently offline
        for item in offlineCBX:
            if item['uniqueId'][17] == "1":
                cp = 0
                searchTerm = item['uniqueId']
            else:
                cp = 1
                searchTerm = item['uniqueId']
                searchTerm = searchTerm[:17] + "1"

            foundRow = searchXL(StatusWB, searchTerm, cpmidColumn, "col")[0]
            #print("XX ", foundRow)
            if foundRow != "notFound":
                StatusWB.cell(row=foundRow, column=TodayColumn).value = "no / offline"
                #StatusWB.cell(row=foundRow, column=cpColumn[cp]).value = "x"
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'Gefunden'
            else:
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'nicht Gefunden'

            TodayFehlerWB.cell(row=todayCbxCounter, column=1).value = item['uniqueId']
            TodayFehlerWB.cell(row=todayCbxCounter, column=3).value = item['masterData']['chargePointName']
            TodayFehlerWB.cell(row=todayCbxCounter, column=4).value = 'offline'
            todayCbxCounter = todayCbxCounter + 1

######################################################################## Fill in all destinations that currently have a failure
        for item in fehlerCBX:

            if item['uniqueId'][17] == "1":
                cp = 0
                searchTerm = item['uniqueId']
            else:
                cp = 1
                searchTerm = item['uniqueId']
                searchTerm = searchTerm[:17] + "1"

            print(searchTerm)
            foundRow = searchXL(StatusWB, searchTerm, cpmidColumn, "col")[0]
            print("Fehlerhafter Standort in Reihe: ", foundRow)
            print(item['ErrorMessage'])

            if foundRow != "notFound":
                cpNo = StatusWB.cell(row=foundRow, column=cpNoColumn).value
                Status = defineFullOrPart(cpidLookup, searchTerm[:12], cpNo)
                print("Fehlerhafter Standort in Reihe: ", foundRow, "mit Status", Status)
                StatusWB.cell(row=foundRow, column=TodayColumn).value = str(Status)
                StatusWB.cell(row=foundRow, column=TodayColumn).font = Font(color="FF0000")
                #StatusWB.cell(row=foundRow, column=cpColumn[cp]).value = "x"
                if StatusWB.cell(row=foundRow, column=TodayColumn-1).value != "j":
                    print(item['ErrorMessage'])
                    #StatusWB.cell(row=foundRow, column=9).value = item['ErrorMessage']
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'Gefunden'
                TodayFehlerWB.cell(row=todayCbxCounter, column=5).value = item['ErrorMessage']
            else:
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'nicht Gefunden'
                TodayFehlerWB.cell(row=todayCbxCounter, column=5).value = item['ErrorMessage']
            TodayFehlerWB.cell(row=todayCbxCounter, column=1).value = item['uniqueId']
            TodayFehlerWB.cell(row=todayCbxCounter, column=3).value = item['chargePointName']
            TodayFehlerWB.cell(row=todayCbxCounter, column=4).value = 'Fehler'
            todayCbxCounter = todayCbxCounter + 1
            wb.save(xlsxPfad)
        i = 1
        # while NIBfehlerWB.cell(row=i, column=1).value != None:
        #     searchTerm = NIBfehlerWB.cell(row=i, column=1).value
        #     foundRow = searchXL(StatusWB, searchTerm, 1, "col")[0]
        #     if foundRow != "notFound":
        #         StatusWB.cell(row=foundRow, column=TodayColumn).value = "j"
        #     i = i + 1


        todayCbxCounter = TodayRow + 1



        while StatusWB.cell(row=todayCbxCounter, column=locationCol).value != None:
            Value = StatusWB.cell(row=todayCbxCounter, column=TodayColumn).value
            ValueM1 = StatusWB.cell(row=todayCbxCounter, column=TodayColumn-1).value
            print("commissionColumn:", str(commissionColumn),"todayCBXCounter", str(todayCbxCounter))
            if Value == None and StatusWB.cell(row=todayCbxCounter, column=commissionColumn).value == "X":
                print(todayCbxCounter)
                StatusWB.cell(row=todayCbxCounter, column=TodayColumn).value = "yes"

            if StatusWB.cell(row=todayCbxCounter, column=TodayColumn).value != ValueM1:
                StatusWB.cell(row=todayCbxCounter, column=cpmidColumn).fill = my_fill
                StatusWB.cell(row=todayCbxCounter, column=cpmidColumn).fill = my_fill
            else:
                StatusWB.cell(row=todayCbxCounter, column=cpmidColumn).fill = no_fill
            changeVal = updateChangeFormula(StatusWB.cell(row=todayCbxCounter, column=TodayColumn-1).coordinate, StatusWB.cell(row=todayCbxCounter, column=TodayColumn).coordinate)
            StatusWB.cell(row=todayCbxCounter, column=changeCol).value = changeVal
            todayCbxCounter = todayCbxCounter + 1


    else:
        print("Es konnte keine Spalte mit dem heutigen Datum gefunden werden. Prüfen Sie die Excel-datei.")

    #xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Downloads\\Aktuelle_IBN_Kopie2.xlsx"
    wb.save(xlsxPfad)



if __name__ == "__main__":
    UserName = os.getlogin()
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Downloads\\IBN_Tracking.xlsx"

    #xlsxPfadFeedback = "C:\\Users\\" + str(UserName) + "\\Desktop\\TrackingFeedback.xlsx"

    f = open("C:/Users/AJ2MSGR/PycharmProjects/WebScrapingForDummies/A2WorkingSkrips/DataFiles/fehlerstandorteStatus.text", 'r')
    fehlerCBX = json.load(f)
    # for element in fehlerCBX:
    #     for item in element:
    #         print(item)
    #
    f = open("C:/Users/AJ2MSGR/PycharmProjects/WebScrapingForDummies/A2WorkingSkrips/DataFiles/offlinestandorte.text", 'r')
    offlineCBX = json.load(f)
    # for element in offlineCBX:
    #     print(element)
    #
    WriteStatusToXL(xlsxPfad, offlineCBX, fehlerCBX)
    #
    # os.startfile(xlsxPfad)
