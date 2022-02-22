from NavigateInExcel import searchXL
import json
from openpyxl import Workbook,load_workbook
import os

def CombineCpSeriealNr():
    f = open("UsableDestinationsDaily.txt", 'r')
    UsableDestinationsCPs = json.load(f)
    wb = load_workbook(filename=r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V8Py.xlsx')

    DataTbl = wb['LSProduktion']# Ursprung der Daten
    ZielSeite = wb['ChargingData']#Ziel der Daten

    IdentificationOrigin = searchXL(DataTbl, "Seriennummer")[1]
    DatenSpalteOrigin = searchXL(DataTbl, "Ladekabel Lieferant Seriennummer")[1]

    SuchSpalteEintrage = searchXL(ZielSeite, "Seriennummer")[1]
    EintragenSeriennummerSpalte = searchXL(ZielSeite, "SeriennummerLadesaule")[1]

    idx = 0 #Anzahl Standorte gefunden anhand des Standort
    AnzlGefunden = 0 #Anzahl Standorte gefunden anhand der ChargePointId

    while
        Status = None
        #Suche der Seriennummer anhand des Chargepoint Standortes (aktuell nur anhand der Stadt)
        Reihe = searchXL(DataTbl, ChargePoint['address']['city'], IdentificationOrigin, "col", SearchBegin)[0]
        if Reihe != "notFound":
            Adresse = DataTbl.cell(row=Reihe, column=IdentificationTarget).value
            Seriennummer = DataTbl.cell(row=Reihe, column=IdentificationOrigin).value
            #Suche des Seriennummereintragungsort anhand der ChargepointID
            ReiheEintragen = searchXL(ZielSeite, ChargePoint['uniqueId'], 1, "col")[0]
            if ReiheEintragen != "notFound":
                Status = "gefunden"
                ZielSeite.cell(row=ReiheEintragen, column=EintragenSeriennummerSpalte).value = Seriennummer
                DataTbl.cell(row=ReiheEintragen, column=FeedbackSpalte).value = Status
                AnzlGefunden = AnzlGefunden+1
            else:
                Status = "Nicht gefunden"
                DataTbl.cell(row=Reihe, column=FeedbackSpalte).value = str("Wurde nicht gefunden")

            print(ChargePoint['uniqueId'], ChargePoint['address']['city'], "Seriennummer:", Seriennummer, " Strasse:", Adresse, " Status:", Status  )
            ReiheM1 = Reihe
            StadtM1 = ChargePoint['address']['city']

            idx = idx+1
        SearchBegin = 1
    print(idx)

    wb.save(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V8Py.xlsx')

if __name__ == "__main__":
    CombineCpSeriealNr()
    os.startfile(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V8Py.xlsx')