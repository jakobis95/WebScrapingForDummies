import json
from Runner import refreshT
from openpyxl import Workbook,load_workbook
import requests
import time
def BackendRequestTemplate(atoken, url, s):
    headers = {
        'authority': 'api.chargepoint-management.com',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Authorization': atoken
    }

    p = s.get(url, headers=headers,  verify=False)
    #print(p)
    #print("return P.TEXT")
    return p

def auswertungBackenddaten(data, atoken, s):
    uniqueIdlcp = ""
    humidity = "first"
    print("Auswertung")
    i = 1
    wb = load_workbook(filename='PythonZuExcel.xlsx')
    Tabellenblatt = wb.active
    obj = data.json()
    print("alle standorte:::::::::")
    for elements in obj['content']:
        if str(elements['uniqueId'])[13:15] == "CP" and int(elements['masterData']['chargingFacilities'][0]['power']) < 350 and str(elements['firmwareVersion']) != "":
            if str(elements['uniqueId'])[:12] == uniqueIdlcp[:12]:
                Tabellenblatt.cell(row=i-1, column=6).value = "CPN012"
            else:
                #print("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:13] + "LMS01?lmsGlobalId=00000000000e0005010f")
                url = "https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:13] + "LMS01?lmsGlobalId=00000000000e0005010f"
                measurement = BackendRequestTemplate(atoken, url, s)
                measurementJ = measurement.json()
                content = measurementJ['message']
                if content == None:
                    humidity = 255
                else:
                    #print(content['idents'][14]['value'])
                    if(content['idents'][14]['value'] == "Closed"):
                        humidity = 999
                    else:
                        humidity = int(content['idents'][14]['value'])
                    #l = 0
                    #for dataM in content['idents']:
                     #   print(str(dataM) + ':' + str(l))
                      #  l = l +1
                #print(elements)
           # print(str(elements['masterData']['chargingFacilities'][0]['power']) + ":" + str(elements['uniqueId']) + ":" + str(elements['masterData']['chargePointName']) + " : " + str(elements['firmwareVersion']))
                Tabellenblatt.cell(row=i, column=1).value = str(elements['uniqueId'])[:2]
                Tabellenblatt.cell(row=i, column=2).value = str(elements['masterData']['chargingFacilities'][0]['power'])
                Tabellenblatt.cell(row=i, column=3).value = str(elements['uniqueId'])
                Tabellenblatt.cell(row=i, column=4).value = str(elements['masterData']['chargePointName'])
                Tabellenblatt.cell(row=i, column=5).value = str(elements['firmwareVersion'])
                Tabellenblatt.cell(row=i, column=6).value = str(elements['uniqueId'])[13:]
                Tabellenblatt.cell(row=i, column=7).value = humidity

                uniqueIdlcp = str(elements['uniqueId'])
                i = i+1

    wb.save('PythonZuExcel.xlsx')

if __name__ == "__main__":
    i = 0
    url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC&status=ACTIVE&status=FAULTED&status=INACTIVE"

    s = requests.session()
    refreshT(s)
    with open('token2.txt', 'r') as jsonf:
        data = json.load(jsonf)
        print("vergleich")
        print(data['refresh_token'])
    atoken = 'Bearer ' + data['access_token']
    data = BackendRequestTemplate(atoken, url, s)
    auswertungBackenddaten(data, atoken, s)