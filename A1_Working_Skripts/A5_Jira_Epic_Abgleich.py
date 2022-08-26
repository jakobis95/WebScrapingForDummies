
import os
from openpyxl import load_workbook
from A2_Working_Support_Functions.NavigateInExcel import searchXL

def update_HVAC_Status(jira_Worksheet, destination_Worksheet, index_column, destination_update_column, destination_index_column, type):
    i = 1
    coordinates = searchXL(jira_Worksheet, index_column, begin=1, rowcol="row", krit=True)
    index_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Summary", begin=1, rowcol="row")
    summary_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Issue Type", begin=1, rowcol="row")
    issue_column = coordinates[1]
    coordinates = searchXL(jira_Worksheet, "Epic Link", begin=1, rowcol="row", krit=True)
    epic_column = coordinates[1]
    coordinates = searchXL(destination_Worksheet, destination_update_column)
    destination_update_column = coordinates[1]
    coordinates = searchXL(destination_Worksheet, destination_index_column)
    destination_index_column = coordinates[1]
    print(type, "###############################")

    i = 2
    missing_Epics = []
    while jira_Worksheet.cell(row=i, column=index_column).value != None:
        index = jira_Worksheet.cell(row=i, column=index_column).value
        coordinates = searchXL(destination_Worksheet, index, begin=destination_index_column, rowcol="col")
        index_row = coordinates[0]
        if index_row != "notFound":
            print("Yes")
            #destination_Worksheet.cell(row=index_row, column=destination_update_column).value = "Ja"
        else:
            string = "Index", jira_Worksheet.cell(row=i, column=index_column).value, "Standort: ", jira_Worksheet.cell(row=i, column=summary_column).value
            missing_Epics.append(string)
        i = i + 1
    for epic in missing_Epics:
        print(epic)

def isEpicLocationMatched(MasterWorksheetPath, destinationWorksheetPath):
    wb_today = load_workbook(filename=destinationWorksheetPath)
    destinationWorksheet = wb_today["STATUS"]
    wb_master = load_workbook(filename=MasterWorksheetPath)
    MasterWorksheet = wb_master["STATUS"]
    coordinates = searchXL(MasterWorksheet, "Jira Epic", begin=1, rowcol="row", krit=True)
    ref_epic_column = coordinates[1]
    coordinates = searchXL(MasterWorksheet, "Location", begin=1, rowcol="row", krit=True)
    ref_name_column = coordinates[1]
    coordinates = searchXL(MasterWorksheet, "Dealer Number", begin=1, rowcol="row", krit=True)
    ref_dealer_column = coordinates[1]
    coordinates = searchXL(destinationWorksheet, "Jira Epic", begin=1, rowcol="row", krit=True)
    epic_column = coordinates[1]
    coordinates = searchXL(destinationWorksheet, "Location", begin=1, rowcol="row", krit=True)
    name_column = coordinates[1]
    coordinates = searchXL(destinationWorksheet, "Dealer Number", begin=1, rowcol="row", krit=True)
    dealer_column = coordinates[1]

    print("######### Starte Abgleich ###########")

    i = coordinates[0] + 1
    missing_Epics = []
    notFoundLocation = []
    while MasterWorksheet.cell(row=i, column=ref_name_column).value != None:
        #ref_Epic = MasterWorksheet.cell(row=i, column=ref_epic_column).value
        #if ref_Epic != None:
        ref_Epic = MasterWorksheet.cell(row=i, column=ref_epic_column).value
        ref_Name = MasterWorksheet.cell(row=i, column=ref_name_column).value
        ref_Dealernumber = MasterWorksheet.cell(row=i, column=ref_dealer_column).value

        coordinates = searchXL(destinationWorksheet, ref_Name, begin=epic_column, rowcol="col")
        index_row = coordinates[0]

        feedback = []
        if index_row != "notFound":
            feedback.append(True)
            comp_Epic = destinationWorksheet.cell(row=index_row, column=epic_column).value
            comp_Name = destinationWorksheet.cell(row=index_row, column=name_column).value
            comp_Dealernumber = destinationWorksheet.cell(row=index_row, column=dealer_column).value

            if ref_Epic == comp_Epic:
                feedback.append(True)
            else:
                feedback.append(False)

            if ref_Name == comp_Name:
                feedback.append(True)
            else:
                feedback.append(False)

            if ref_Dealernumber == comp_Dealernumber:
                feedback.append(True)
            else:
                feedback.append(False)

        else:
            feedback.append(False)
            feedback.append(False)
            feedback.append(False)
            feedback.append(False)

        for state in feedback:
            if state != True:
                Str = "[ Epic of " + str(ref_Name) + " found: " + str(feedback[0]) + "] [Epic correct: " + str(feedback[1]) + "] [Location Name correct: " + str(feedback[2]) + "] [Dealernumber correct: " + str(feedback[3]) +"]"
                if feedback[0]:
                    missing_Epics.append(Str)
                else:
                    notFoundLocation.append(Str)
                #print("Coordinaten Epic: ", coordinates[0], coordinates[1])
                break

        i = i + 1
    print("######### Abgleich Beendet ###########")
    print("\nEpic or Dealernumber dont Match:")
    for epic in missing_Epics:
        print(epic)
    print("\nLocation cant be found:")
    for epic in notFoundLocation:
        print(epic)

if __name__ == "__main__":
#Todo Alle leeren Felder mit Nein befüllen
    UserName = os.getlogin()
    jira_xlsx_path_HVAC = "C:\\Users\\" + str(UserName) + "\\Downloads\\JiraEpics.xlsx"
    master_xlsx_path = "C:\\Users\\" + str(UserName) + "\\Downloads\\DO_NOT_TOUCHE_JIRA_MASTERLISTE.xlsx"
    today_ibn_path = "C:\\Users\\" + str(UserName) + "\\Downloads\\IBN_Tracking.xlsx"
#Worksheets
    wb_today = load_workbook(filename=today_ibn_path)
    status_ws = wb_today["STATUS"]
    wb_master = load_workbook(filename=master_xlsx_path)
    master_status_ws = wb_master["STATUS"]
    wb_jira = load_workbook(filename=jira_xlsx_path_HVAC)
    jira_ws_HVAC = wb_jira["Sheet1"]
#Skriptaufrufe
    #update_HVAC_Status(jira_ws_HVAC, status_ws, "Key", "HVAC Maßnahme\numgesetzt (WMQ1)", "Jira Epic", "HVAC")
    isEpicLocationMatched(master_xlsx_path, today_ibn_path)
    os.startfile(master_xlsx_path)
    os.startfile(today_ibn_path)