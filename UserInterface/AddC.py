import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from copy import copy

def addColToRef(ref, additionalCols):
    refStart = ref.split(":")[0]
    refEnd = ref.split(":")[1]
    refRow,refCol = coordinate_to_tuple(refEnd)
    newCol = get_column_letter(refCol + additionalCols)
    newRef = refStart + ":" + newCol + str(refRow)
    return newRef


def addShit(ws):
    ws.append(["Fruit", "2011", "2012", "2013", "2014","2011", "2012", "2013", "2014"])

    data = [
        ['Apples', 10000, 5000, 8000, 6000],
        ['Pears', 2000, 3000, 4000, 5000],
        ['Bananas', 6000, 6000, 6500, 6000],
        ['Oranges', 500, 300, 200, 700],
        ['Apples', 10000, 5000, 8000, 6000],
        ['Pears', 2000, 3000, 4000, 5000],
        ['Bananas', 6000, 6000, 6500, 6000],
        ['Oranges', 500, 300, 200, 700],
    ]
    for row in data:
        ws.append(row)
    tab = Table(displayName="Table1", ref="A1:E5")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def addCol(ws, colNr ,headerRow , headerVal):
    ws.insert_cols(colNr)
    ws.cell(row=headerRow, column=colNr).value = headerVal
    #Adds correct color to Today column
    ws.cell(row=headerRow, column=colNr).fill = copy(ws.cell(row=headerRow, column=colNr-1).fill)
    ws.cell(row=headerRow, column=colNr).alignment = copy(ws.cell(row=headerRow, column=colNr-1).alignment)

def update(xlsxPfad, sheetName):
    check = False
    while not check:
        try:
            wb = load_workbook(filename=xlsxPfad)
            check = True
        except:
            input("Excel konnte nicht gupdated werden, stellen Sie sich das die Datei geschlossen ist\n um Update zu wiederholen y drücken")
    ws = wb[sheetName]
    wb.save(xlsxPfad)
    return wb,ws

def replaceTab(ws, ref, displayName):
    #copys autofilter
    autofilter = ws.tables[displayName].autoFilter
    #copy current style
    style = ws.tables[displayName].tableStyleInfo
    #create new enlarged table
    entab = Table(displayName=displayName, ref=ref)
    entab.tableStyleInfo = style
    entab.autoFilter = autofilter
    #change current table to new enlarged tablead
    ws.tables[displayName] = entab




"""
Wenn etwas in einer bestehenden Tabelle geändert werden muss danach auch die Tabelle erneuert werden soll wird die xlsx nicht zu oeffnen sein
bisher ist mir das beim hinzufügen von spalten und dem Verändern der Spalten Überschrift aufgefallen
beim verändern des alignment der Überschriften und des Fill
"""
if __name__ == "__main__":
    UserName = os.getlogin()
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Downloads\\book1.xlsx"
    wb = load_workbook(filename=xlsxPfad)
    ws = wb["Sheet1"]

    addShit(ws)
    wb.save(xlsxPfad)
    wb, ws = update(xlsxPfad, "Sheet1")
    #add column in column 2
    addCol(ws, 2, 1, "NewCol1")
    # update table size
    replaceTab(ws, "A1:F5", "Table1")
    wb.save(xlsxPfad)
    #update workbook and worksheet
    wb, ws = update(xlsxPfad, "Sheet1")
    # add column in column 2
    addCol(ws, 2, 1, "NewCol2")
    #update table size
    replaceTab(ws, "A1:G5", "Table1")
    wb.save(xlsxPfad)
    os.startfile(xlsxPfad)
