from openpyxl import load_workbook, styles
import os
def takeRow(elem):
    return elem[1]

def main(xlsxPfad):

    wb = load_workbook(filename=xlsxPfad, keep_vba=True)
    IBN_S = wb.worksheets[0]
    maxRow = IBN_S.max_row
    maxCol = IBN_S.max_column
    No = 0
    dataSheet = []
    cell = []
    for column in range(1, maxCol, 1):
        for row in range(1, maxRow, 1):
            if IBN_S.cell(row=row, column=column).value:
                No = No + 1
                cell = (column, row, IBN_S.cell(row=row, column=column).value)
                dataSheet.append(cell)
                #print("Number",No,": Cor(",column, ";", row,")")
                #print(IBN_S.cell(row=row, column=column).value)
    dataSheet.sort(key=takeRow)
    firstRow = dataSheet[0][1]
    for value in dataSheet:
        if firstRow != value[1]:
            firstRow = value[1]
            begin = "\n"
        else:
            begin = "; "
        print(begin, value[2], end="")
    #wb.save(xlsxPfad)
if __name__ == "__main__":
    UserName = os.getlogin()
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Documents\\teesside cbx commissioning.xlsm"
    #xlsxPfad = "C:/Users/AJ2MSGR/OneDrivea - Dr.Ing.h.c.F.Porsche AG/IBN_Kastenfinden.xlsm"
    main(xlsxPfad)