from openpyxl import load_workbook, styles
from xml.dom import minidom
from XMLread import followXMLPath, ridOf
import os
from readXlIbn import returnPDFdata,takeRow

def getBatteryIdentList(ibnData):
    batteryIdentList = []
    for row in range(63, 70, 2):
        for batteryIdent in ibnData[row]:
            batteryIdentList.append(str(batteryIdent))
    for row in range(73, 80, 2):
        for batteryIdent in ibnData[row]:
            batteryIdentList.append(str(batteryIdent))
    i = 0
    for bat in batteryIdentList:
        print(i,bat)
        i = i + 1
    return batteryIdentList

def createList(string):
    if string != None:
        return string.split()

if __name__ == "__main__":
    UserName = os.getlogin()
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Documents\\Annex_G_ Commissioning_Report_KLL_Template_en.xlsm"
    ibnData = returnPDFdata(xlsxPfad)


    xlsxPfad="C:\\Users\\" + str(UserName) + "\\Documents\\Strucktur_Zuordnungstabelle_IBN_Protokolle.xlsx"
    wb = load_workbook(filename=xlsxPfad, keep_vba=True)
    ws = wb.worksheets[0]

    xmlSheetPath =  "C:\\Users\\AJ2MSGR\\Documents\\Template_ASNAST_XML_ChargeBox.xml"
    xmldoc = minidom.parse(xmlSheetPath)
    startNode = xmldoc.getElementsByTagName("ZSB_PCB")
    path = ["ITEM"]
    startNode = startNode[0]
    parent = followXMLPath(startNode, path)
    itemNode = parent[1]
    i = 2
    new = ws.cell(row=i, column=1).value
    old = ws.cell(row=i, column=1).value
    nodeId = 1
    batteryIdentList = getBatteryIdentList(ibnData)
    for i in range(2, 28, 1):
        new = ws.cell(row=i, column=1).value
        if new != old:
            nodeId = nodeId + 2
        print(itemNode.childNodes[nodeId].localName)
        if new == "CBX_BATTERY":
            batteryId = 0
            batterys = itemNode.childNodes[nodeId]
            for battery in batterys.childNodes:
                fullType = 0
                print(battery.localName)
                if str(battery.localName) == "BATTERY" :
                    print("BatteryID:",batteryId, "Serienummer", batteryIdentList[batteryId])
                    path = ["IDENT"]
                    batteryIdent = followXMLPath(battery,path)[1]
                    nodeText = xmldoc.createTextNode(batteryIdentList[batteryId])
                    batteryIdent.appendChild(nodeText)
                    batteryId = batteryId + 1
                elif str(battery.localName) == "TYPE":
                    print(battery.localName,ibnData[int(ws.cell(row=i, column=5).value)][int(ws.cell(row=i, column=6).value)])
                    nodeText = xmldoc.createTextNode(str(ibnData[int(ws.cell(row=i, column=5).value)][int(ws.cell(row=i, column=6).value)]).upper())
                    battery.appendChild(nodeText)


            #xmldoc.writexml(open('East_Flanders_IBN.xml', 'w'), indent="  ", addindent="  ", newl='\n')
        elif new == "POWER_SUPPLY":
            drawing = "C:\\Users\\AJ2MSGR\\Documents\\Commissioning Report - PC East-Flanders - 20220718\\xl\drawings\\drawing1.xml"
            sheet = "C:\\Users\\AJ2MSGR\\Documents\\Commissioning Report - PC East-Flanders - 20220718\\xl\\worksheets\\sheet1.xml"
            ctrPropsDir = "C:\\Users\\AJ2MSGR\\Documents\\Commissioning Report - PC East-Flanders - 20220718\\xl\\ctrlProps"
            relSheet = "C:\\Users\\AJ2MSGR\\Documents\\Commissioning Report - PC East-Flanders - 20220718\\xl\\worksheets\\_rels\\sheet1.xml.rels"
            path = createList(ws.cell(row=i, column=3).value)
            print("path", path)
            node = followXMLPath(itemNode.childNodes[nodeId], path)
            nodeData = node[1].firstChild
            print(node[1].localName, ibnData[int(ws.cell(row=i, column=5).value)][int(ws.cell(row=i, column=6).value)])
            nodeText = xmldoc.createTextNode(str(ridOf(sheet, relSheet, drawing, ctrPropsDir)))
            node[1].appendChild(nodeText)
        else:
            path = createList(ws.cell(row=i, column=3).value)
            print("path", path)
            node = followXMLPath(itemNode.childNodes[nodeId],path)
            nodeData = node[1].firstChild
            print(node[1].localName, ibnData[int(ws.cell(row=i, column=5).value)][int(ws.cell(row=i, column=6).value)])
            nodeText = xmldoc.createTextNode(str(ibnData[int(ws.cell(row=i, column=5).value)][int(ws.cell(row=i, column=6).value)]))
            node[1].appendChild(nodeText)
        old = ws.cell(row=i, column=1).value
    xmldoc.writexml(open('Romania_IBN.xml', 'w'), indent="  ", addindent="  ", newl='\n')
    #todo Die Netzteil leistung wird irgendwie nicht richtig eingelesen.