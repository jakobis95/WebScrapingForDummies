import requests
import json
import time
from openpyxl import Workbook,load_workbook
from Runner import refreshT


def loadalldata(atoken, chargePointsIds, cpoIds, s):
    #url = "https://api.chargepoint-management.com/status/connectionStatusList"
    url = "https://api.chargepoint-management.com/chargepoint/v1/transaction/list/filter?page=0&size=10&sort=startDate,desc"
    headers = {
        'authority': 'api.chargepoint-management.com',
        'method' :  'POST',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Authorization': atoken,
        'content-length' : '1594',
        'content-type': 'application/json',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36'
    }
    #payload = ["ES9110000150_CPN01","GB9110000272_LMS01","GB9110000272_CPN01","GB9110000272_CPN02","US9110000298_LMS01","US9110000298_CPN01"]
    payload = {"cpoIds":[2,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,64,65,66,67,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208,209,211,212,213,214,215,217,218,219,220,221,222,223,224,225,226,227,228,229,230,231,232,233,234,235,236,237,238,239,240,241,242,243,244,245,246,247,248,249,250,251,252,253,254,255,256,257,258,259,260,261,262,263,264,265,266,267,268,269,270,271,272,273,274,275,276,277,278,279,280,281,282,283,284,285,286,287,288,289,290,291,292,293,294,295,296,297,298,299,301,302,303,304,305,306,307,308,309,310,311,312,313,314,315,316,317,318,319,320,321,322,323,324,325,326,327,328,329,330,331,332,333,334,335,336,337,338,339,340,341,342,343,344,345,346,347,348,349,350,351,352,353,354,355,356,357,358,359,360,361,362,363,364,365,366,367,368,369,370,371,373,374,375,376,378,379,380,381,382,383,384,385,386,387,388,389,390,391,392,393,394,395,396,397,398,399,400,401,402,403,404,405,406,407,408,409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425],"chargePointIds":[]}
    p = s.post(url, headers=headers, verify=False, json=payload)
    print(p.json())
    print(p)
    with open('TransactionData.json', 'w') as f:
        f.write(p.text)
    return p.json()

def Create_Database_From_Json():
    wb = Workbook()
    ws = wb.worksheets[0]
    r = 1
    col = 1
    with open('TransactionData.json', 'r') as f:
        obj = json.load(f)
        for cpTransactions in obj['content']:
            if r != 1:
                for index in cpTransactions:
                    print(cpTransactions[index])
                    ws.cell(row=r, column=col).value = str(cpTransactions[index])
                    col = col + 1
                col = 1
            else:
                for index in cpTransactions:
                    print("first row, header is created for column %d", col)
                    ws.cell(row=r, column=col).value = str(index)
                    col = col + 1
                col = 1
                r = r + 1
                for index in cpTransactions:
                    print(cpTransactions[index])
                    ws.cell(row=r, column=col).value = str(cpTransactions[index])
                    col = col + 1
                col = 1

            col = 1
            r = r + 1
    wb.save('TransactionData.xlsx')
def Get_cpoIds():
    #Copy of the cpoIds payload from the Backend
    cpoIds = {2,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,64,65,66,67,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208,209,211,212,213,214,215,217,218,219,220,221,222,223,224,225,226,227,228,229,230,231,232,233,234,235,236,237,238,239,240,241,242,243,244,245,246,247,248,249,250,251,252,253,254,255,256,257,258,259,260,261,262,263,264,265,266,267,268,269,270,271,272,273,274,275,276,277,278,279,280,281,282,283,284,285,286,287,288,289,290,291,292,293,294,295,296,297,298,299,301,302,303,304,305,306,307,308,309,310,311,312,313,314,315,316,317,318,319,320,321,322,323,324,325,326,327,328,329,330,331,332,333,334,335,336,337,338,339,340,341,342,343,344,345,346,347,348,349,350,351,352,353,354,355,356,357,358,359,360,361,362,363,364,365,366,367,368,369,370,371,373,374,375,376,378,379,380,381,382,383,384,385,386,387,388,389,390,391,392,393,394,395,396,397,398,399,400,401,402,403,404,405,406,407,408,409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425}
    print(cpoIds)
    return cpoIds

if __name__ == '__main__':
    # s = requests.session()
    # refreshT(s)
    # with open('token2.txt', 'r') as jsonf:
    #     data = json.load(jsonf)
    #     print("vergleich")
    #     print( data['refresh_token'])
    # atoken = 'Bearer ' + data['access_token']
    # chargePointsIds = []
    # cpoIds = Get_cpoIds()
    # loadalldata(atoken, chargePointsIds, cpoIds, s)
    Create_Database_From_Json()