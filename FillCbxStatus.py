import json
import os
from openpyxl import Workbook,load_workbook, styles
from NavigateInExcel import searchXL
from datetime import datetime

def WriteStatusToXL(xlsxPfad, offlineCBX, fehlerCBX):
    cp = 0
    wb = load_workbook(filename=xlsxPfad)
    todayCbxCounter = 1
    StatusWB = wb["cbxStatusListe"]
    TodayFehlerWB = wb["overviewToday"]
    NIBfehlerWB = wb["NIBfehler"]
    NIBofflineWB = wb["NIBoffline"]
    LastRow = StatusWB.max_row
    TodayColumnString = "Fehlerhaft am " + datetime.today().strftime('%d.%m.%Y')
    #TodayColumnString = "Fehlerhaft am 28.01.2022"
    #finds Column with the date of today
    TodayCell = searchXL(StatusWB, TodayColumnString)
    TodayColumn = TodayCell[1]
    TodayRow = TodayCell[0]
    #finds Column to display if cp1 and/or cp2 has a problem
    cp1Column = searchXL(StatusWB, "1")[1]
    cp2Column = searchXL(StatusWB, "2")[1]
    cpColumn = [cp1Column,cp2Column]

    if TodayColumn != "notFound":
        print("XX today Column is:", TodayColumn)
        #Fill in all destinations that are currently offline
        for item in offlineCBX:
            if item['uniqueId'][17] == "1":
                cp = 0
                searchTerm = item['uniqueId']
            else:
                cp = 1
                searchTerm = item['uniqueId']
                searchTerm = searchTerm[:17] + "1"

            foundRow = searchXL(StatusWB, searchTerm, 1, "col")[0]
            print("XX ", foundRow)
            if foundRow != "notFound":
                StatusWB.cell(row=foundRow, column=TodayColumn).value = "offline"
                StatusWB.cell(row=foundRow, column=cpColumn[cp]).value = "x"
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'Gefunden'
            else:
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'nicht Gefunden'

            TodayFehlerWB.cell(row=todayCbxCounter, column=1).value = item['uniqueId']
            TodayFehlerWB.cell(row=todayCbxCounter, column=3).value = item['masterData']['chargePointName']
            TodayFehlerWB.cell(row=todayCbxCounter, column=4).value = 'offline'
            todayCbxCounter = todayCbxCounter + 1

        # Fill in all destinations that currently have a failure
        for item in fehlerCBX:

            if item['uniqueId'][17] == "1":
                cp = 0
                searchTerm = item['uniqueId']
            else:
                cp = 1
                searchTerm = item['uniqueId']
                searchTerm = searchTerm[:17] + "1"

            print(searchTerm)
            foundRow = searchXL(StatusWB, searchTerm, 1, "col")[0]
            print("XX ", foundRow)

            if foundRow != "notFound":
                StatusWB.cell(row=foundRow, column=TodayColumn).value = "j"
                StatusWB.cell(row=foundRow, column=cpColumn[cp]).value = "x"
                if StatusWB.cell(row=foundRow, column=TodayColumn-1).value != "j":
                    StatusWB.cell(row=foundRow, column=9).value = item['ErrorMessage']
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'Gefunden'
            else:
                TodayFehlerWB.cell(row=todayCbxCounter, column=2).value = 'nicht Gefunden'

            TodayFehlerWB.cell(row=todayCbxCounter, column=1).value = item['uniqueId']
            TodayFehlerWB.cell(row=todayCbxCounter, column=3).value = item['chargePointName']
            TodayFehlerWB.cell(row=todayCbxCounter, column=4).value = 'Fehler'
            todayCbxCounter = todayCbxCounter + 1

        i = 1
        while NIBfehlerWB.cell(row=i, column=1).value != None:
            searchTerm = NIBfehlerWB.cell(row=i, column=1).value
            foundRow = searchXL(StatusWB, searchTerm, 1, "col")[0]
            StatusWB.cell(row=foundRow, column=TodayColumn).value = "j"
            i = i + 1

    # TODO: Gelb maikieren wenn Zustandänderung
        todayCbxCounter = 7
        my_yellow = styles.colors.Color(rgb='ffff00')
        my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
        no_fill = styles.PatternFill(fill_type=None)

        while StatusWB.cell(row=todayCbxCounter, column=1).value != None:
            Value = StatusWB.cell(row=todayCbxCounter, column=TodayColumn).value
            ValueM1 = StatusWB.cell(row=todayCbxCounter, column=TodayColumn-1).value
            if Value == None :
                print(todayCbxCounter)
                StatusWB.cell(row=todayCbxCounter, column=TodayColumn).value = "n"

            if StatusWB.cell(row=todayCbxCounter, column=TodayColumn).value != ValueM1:
                StatusWB.cell(row=todayCbxCounter, column=3).fill = my_fill
            else:
                StatusWB.cell(row=todayCbxCounter, column=3).fill = no_fill
            todayCbxCounter = todayCbxCounter + 1

    else:
        print("Es konnte kein Spalte mit dem heutigen Datum gefunden werden. Prüfen Sie die Excel-datei.")
    # print(searchXL(StatusWB, "Hello was geht"))
    wb.save(xlsxPfad)

#def WriteStatusToXL(offlineCBX, fehlerCBX, xlsxPfad):

if __name__ == "__main__":
    xlsxPfad = r"C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\LE_Failure_Analysis\CBX_Fehlerliste_AutoPyTesting.xlsx"

    #WriteStatusToXL(offlineCBX, fehlerCBX, xlsxPfad)
    #WriteStatusToXL( xlsxPfad, offlineCBX, offlineCBX)
    f = open("fehlerstandorteStatus.text", 'r')
    fehlerCBX = json.load(f)
    for element in fehlerCBX:
        for item in element:
            print(item)

    f = open("offlinestandorte.text", 'r')
    offlineCBX = json.load(f)
    for element in offlineCBX:
        print(element)

    WriteStatusToXL(xlsxPfad, offlineCBX, fehlerCBX)
    xlsxPfad = r"C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\LE_Failure_Analysis\CBX_Fehlerliste_AutoPyTesting.xlsx"
    #os.system('start  "excel" "C:\\Users\\FO4A5OY\\OneDrive - Dr. Ing. h.c. F. Porsche AG\\LE_Failure_Analysis\\CBX_Fehlerliste_AutoPyTesting.xlsx"')
    os.startfile(xlsxPfad)
