from NavigateInExcel import searchXL
import json
from openpyxl import Workbook,load_workbook
import os

def CombineCpSeriealNr():
    f = open("UsableDestinationsDaily.txt", 'r')
    UsableDestinationsCPs = json.load(f)
    wb = load_workbook(filename=r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V8Py.xlsx')

    SeriennummernTbl = wb['Seriennummern']
    CpSerKombiTbl = wb['ChargingData']

    SeriennummerSpalte = searchXL(SeriennummernTbl, "Seriennummer")[1]
    StadtSpalte = searchXL(SeriennummernTbl, "Ort")[1]
    StrasseSpalte = searchXL(SeriennummernTbl, "Adresse")[1]
    FeedbackSpalte = searchXL(SeriennummernTbl, "Feedback")[1]

    SuchSpalteEintrage = searchXL(CpSerKombiTbl, "ChargePointId")[1]
    EintragenSeriennummerSpalte = searchXL(CpSerKombiTbl, "SeriennummerLadesaule")[1]
    ReiheM1 = 1
    StadtM1 = 1

    idx = 0 #Anzahl Standorte gefunden anhand des Standort
    AnzlGefunden = 0 #Anzahl Standorte gefunden anhand der ChargePointId

    for ChargePoint in UsableDestinationsCPs:
        Status = None
        if ChargePoint['address']['city'] != StadtM1:
            SearchBegin = 1
        else:
            SearchBegin = ReiheM1+1
        #Suche der Seriennummer anhand des Chargepoint Standortes (aktuell nur anhand der Stadt)
        Reihe = searchXL(SeriennummernTbl, ChargePoint['address']['city'], StadtSpalte, "col", SearchBegin )[0]
        if Reihe != "notFound":
            Adresse = SeriennummernTbl.cell(row=Reihe, column=StrasseSpalte).value
            Seriennummer = SeriennummernTbl.cell(row=Reihe, column=SeriennummerSpalte).value
            #Suche des Seriennummereintragungsort anhand der ChargepointID
            ReiheEintragen = searchXL(CpSerKombiTbl, ChargePoint['uniqueId'], 1, "col")[0]
            if ReiheEintragen != "notFound":
                Status = "gefunden"
                CpSerKombiTbl.cell(row=ReiheEintragen, column=EintragenSeriennummerSpalte).value = Seriennummer
                SeriennummernTbl.cell(row=ReiheEintragen, column=FeedbackSpalte).value = Status
                AnzlGefunden = AnzlGefunden+1
            else:
                Status = "Nicht gefunden"
                SeriennummernTbl.cell(row=Reihe, column=FeedbackSpalte).value = str("Wurde nicht gefunden")

            print(ChargePoint['uniqueId'], ChargePoint['address']['city'], "Seriennummer:", Seriennummer, " Strasse:", Adresse, " Status:", Status  )
            ReiheM1 = Reihe
            StadtM1 = ChargePoint['address']['city']

            idx = idx+1
        SearchBegin = 1
    print(idx)

    wb.save(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V8Py.xlsx')

if __name__ == "__main__":
    CombineCpSeriealNr()
    os.startfile(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V8Py.xlsx')