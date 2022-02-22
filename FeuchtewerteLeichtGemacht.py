import json
import os
from Runner import refreshT
from openpyxl import Workbook,load_workbook
import requests
from NavigateInExcel import searchXL, conditional_formatting_with_rules
from datetime import datetime
import time
def findTodayCol(worksheet):
    # soll die letzte beschriebene Spalte finden
    LastCol = worksheet.max_column
    print("LastCol ", LastCol)
    i = 0
    ok = False
    while ok != True:
        WasStehtHierDrinne = worksheet.cell(row=1, column=LastCol - i).value
        if WasStehtHierDrinne != None:
            print("Letze Spalte", LastCol - i, " mit Inhalt", WasStehtHierDrinne)
            ok = True
        else:
            print("Letze Spalte hat keinen Inalt counter =", i)
            i = i + 1

    LastCol = LastCol - i
    print("XX Letze Spalte", LastCol, " mit Inhalt", WasStehtHierDrinne)
    return LastCol

def OnlyUsableDestinations(data):
    obj = data.json()
    CPlist = []
    for elements in obj['content']:
        if str(elements['uniqueId'])[13:15] == "CP" and int(elements['masterData']['chargingFacilities'][0]['power']) < 350 and str(elements['firmwareVersion']) != "" and int(elements['masterData']['chargingFacilities'][0]['power']) > 150:
            CPlist.append(elements)

    return CPlist
def BackendRequestTemplate(atoken, url, s):
    headers = {
        'authority': 'api.chargepoint-management.com',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Authorization': atoken
    }

    p = s.get(url, headers=headers,  verify=False)
    #print(p)
    #print("return P.TEXT")
    return p

def auswertungBackenddaten(obj, atoken, s, filepath):
    uniqueIdlcp = ""
    humidity = "999"
    print("Auswertung")
    i = 1
    wb = load_workbook(filename=filepath)
    FeuchteTbl = wb.worksheets[0]
    FeedbackTbl = wb.worksheets[1]

    LastCol = findTodayCol(FeuchteTbl)
    FirmwareCol = searchXL(FeuchteTbl,"Firmware")[1]
    TodayCol = LastCol + 2



    FeuchteTbl.cell(row=1, column=TodayCol).value = "Feuchte " + datetime.today().strftime('%d.%m.%Y %H:%M:%S')


    j = 1
    CPlist = []
    curDateTime = datetime.today()
    file = open("lastHumidityUpdateLog", "r+")
    lastHumidityUpdateLog= file.read()
    if lastHumidityUpdateLog == None:
        print("Kein letztes Update dokumentier")
        file.write(str(curDateTime))
    lastHumidityUpdateTime = datetime.strptime(lastHumidityUpdateLog,'%Y-%m-%d %H:%M:%S.%f')
    timeSinceLastUpdate = curDateTime - lastHumidityUpdateTime
    print("minutes since last Update:", timeSinceLastUpdate)
    #Updates chargebox measurements in Backend
    if timeSinceLastUpdate.total_seconds()/60 <= 30:
        print("Feuchte Werte sind noch aktuell")
    for elements in obj:
        if str(elements['uniqueId'])[13:18] == "CPN01":
            url = "https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:13] + "LMS01/request?lmsGlobalId=0000000000000005010f&force=true"
            update_message = BackendRequestTemplate(atoken, url, s)
            print(str(elements['uniqueId'])[13:18] + " i=" + str(j))
            CPlist.append(elements)
            j= j +1
    lastHumidityUpdateLog = open("lastHumidityUpdateLog", "w")
    lastHumidityUpdateLog.write(str(curDateTime))

    i = 1
    #Pulls chargebox mesurement data from Backend
    for elements in CPlist:
        print(str(elements['uniqueId'])[13:18])
        print(str(elements['uniqueId']))
        url = "https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:13] + "LMS01?lmsGlobalId=00000000000e0005010f"
        measurement = BackendRequestTemplate(atoken, url, s)
        measurementJ = measurement.json()
        content = measurementJ['message']
        if content == None:
            humidity = 255
        else:
            if content['idents'][14]['value'] == "Closed" or content['idents'][14]['value'] == "" :
                print(content['idents'][14]['value'])
                humidity = 999
            elif content['idents'][14]['value'].isnumeric() == True:
                #print(content['idents'][14]['value'])
                humidity = content['idents'][14]['value']
            else:
                print("not numeric" + content['idents'][14]['value'])
                humidity = 777

#searchID and fill in humidity
        Xcp = searchXL(FeuchteTbl, str(elements['uniqueId']))[0]
        if Xcp != "notFound":
            #kleine Sicherheitsfunktion falls Neuestandorte hinzugefügt wurden ohne für diesen Tag den entsprechenden Wert hinzuzufügen
            if FeuchteTbl.cell(row=Xcp, column=TodayCol).value == None:
                FeuchteTbl.cell(row=Xcp, column=TodayCol).value = "0"

            print("feuchte", int(humidity))
            FeuchteTbl.cell(row=Xcp, column=TodayCol).value = int(humidity)
            FeuchteTbl.cell(row=Xcp, column=FirmwareCol).value = str(elements['firmwareVersion'])
            print("Tabellenwert Heute:", FeuchteTbl.cell(row=Xcp, column=TodayCol).value," - ", FeuchteTbl.cell(row=Xcp, column=TodayCol-2).value)
            FeuchteTbl.cell(row=Xcp, column=TodayCol-1).value = FeuchteTbl.cell(row=Xcp, column=TodayCol).value - FeuchteTbl.cell(row=Xcp, column=TodayCol-2).value
            FeedbackMessage = "Found in row: " + str(Xcp)
        else:
            FeedbackMessage = "Not found"

        FeedbackTbl.cell(row=i, column=1).value = str(elements['uniqueId'])[:2]
        FeedbackTbl.cell(row=i, column=2).value = str(elements['masterData']['chargingFacilities'][0]['power'])
        FeedbackTbl.cell(row=i, column=3).value = str(elements['uniqueId'])
        FeedbackTbl.cell(row=i, column=4).value = str(elements['masterData']['chargePointName'])
        FeedbackTbl.cell(row=i, column=5).value = str(elements['firmwareVersion'])
        FeedbackTbl.cell(row=i, column=6).value = str(elements['uniqueId'])[13:]
        FeedbackTbl.cell(row=i, column=7).value = int(humidity)
        FeedbackTbl.cell(row=i, column=8).value = FeedbackMessage
        i = i+1

    conditional_formatting_with_rules(FeuchteTbl, TodayCol)
    wb.save(filepath)

if __name__ == "__main__":
    i = 0
    #url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC&status=ACTIVE&status=FAULTED&status=INACTIVE"
    url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC"
    s = requests.session()
    UserName = os.getlogin()

    filepath = "C:\\Users\\"+ str(UserName) +"\\Dr. Ing. h.c. F. Porsche AG\\Rollout KLL - Task Force HVAC\\Feuchte_Overview_CBX.xlsx"
    #filepath = r'C:\Users\FO4A5OY\Dr. Ing. h.c. F. Porsche AG\Rollout KLL - Task Force HVAC\Feuchte_Overview_CBX.xlsx'
    refreshT(s)
    with open('token2.txt', 'r') as jsonf:
        data = json.load(jsonf)
        print("vergleich")
        print(data['refresh_token'])
    atoken = 'Bearer ' + data['access_token']
    data = BackendRequestTemplate(atoken, url, s)
    data = OnlyUsableDestinations(data)
    #print(data)
    f = open("UsableDestinationsDaily.txt", 'w')
    f.write(json.dumps(data))
    auswertungBackenddaten(data, atoken, s, filepath)

    os.startfile(filepath)
    # wb = load_workbook(filename='PythonZuExcel.xlsx')
    # FeuchteTbl = wb.worksheets[0]
    # LastCol = findTodayCol(FeuchteTbl)
    # for i in range(1,100):
    #     print(FeuchteTbl.cell(row=i, column=LastCol).value)