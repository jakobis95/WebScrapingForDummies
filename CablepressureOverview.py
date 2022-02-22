import json
import os
from Runner import refreshT, BackendRequestTemplate as BackReqTem
from openpyxl import Workbook,load_workbook
import requests
from NavigateInExcel import searchXL, conditional_formatting_with_rules
from datetime import datetime
import time
from FeuchtewerteLeichtGemacht import BackendRequestTemplate, OnlyUsableDestinations
def Update(CP,atoken,s):
    elements = CP

    headers = {
        'authority': 'api.chargepoint-management.com',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Authorization': atoken
    }

    # update Ladestatus
    url = ("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:12] + "_LMS01/request?lmsGlobalId=00000000000100030" +str(elements['uniqueId'])[17] + "0a&force=true")
    # url = ("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:12] + "_LMS01/request?lmsGlobalId=de911000016700030"+ str(elements['uniqueId'])[17] +"0a&force=true")
    s.get(url, headers=headers, verify=False)
    time.sleep(2)
def PullPressure(CP, atoken, s):
    elements = CP

    headers = {
        'authority': 'api.chargepoint-management.com',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Authorization': atoken
    }

    # get mesurements loading control
    url = ("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:12] + "_LMS01?lmsGlobalId=00000000000100030" +str(elements['uniqueId'])[17] + "0a")
    print(url)
    CPdata = s.get(url, headers=headers, verify=False)
    CPdataJ = CPdata.json()
    return CPdataJ
def CpPressureToXl(data, atoken, s):
    wb = load_workbook(filename='PythonZuExcel.xlsx')
    FeuchteTbl = wb.worksheets[0]
    FeedbackTbl = wb.worksheets[1]
    CablePrs1 = searchXL(FeuchteTbl, "cablePressure1")[1]
    CablePrs2 = searchXL(FeuchteTbl, "cablePressure2")[1]
    CablePrs = (CablePrs1, CablePrs2)
    j = 2
    for element in data:
        print(element)
        CPdata = PullPressure(element, atoken, s)
        searchStr = str(element['uniqueId'])[:17] + "1"
        print("searchstr:", searchStr)
        Xcp = searchXL(FeuchteTbl, searchStr)[0]
        if Xcp != "notFound":
            if str(element['uniqueId'])[17] == "1":
                print("CP 1 found")
                i = 0
            elif str(element['uniqueId'])[17] == "2":
                print("CP 2 found")
                i = 1
            else:
                print("error CP ID besides 1 or 2 found")
                break
            print(CPdata)
            if CPdata['message'] != None:
                FeuchteTbl.cell(row=Xcp, column=CablePrs[i]).value = CPdata['message']['idents'][16]['value']
            else:
                FeuchteTbl.cell(row=Xcp, column=CablePrs[i]).value = "None"
            FeedbackMessage = "Found in Row:" + str(Xcp)
        else:
            FeedbackMessage = "NotFound"
        FeedbackTbl.cell(row=j, column=1).value = str(element['uniqueId'])[:2]
        FeedbackTbl.cell(row=j, column=2).value = str(element['masterData']['chargingFacilities'][0]['power'])
        FeedbackTbl.cell(row=j, column=3).value = str(element['uniqueId'])
        FeedbackTbl.cell(row=j, column=4).value = str(element['masterData']['chargePointName'])
        FeedbackTbl.cell(row=j, column=5).value = str(element['firmwareVersion'])
        FeedbackTbl.cell(row=j, column=6).value = str(element['uniqueId'])[13:]
        FeedbackTbl.cell(row=j, column=7).value = FeedbackMessage
        j = j + 1
        if j % 10 == 0:
            atoken = GetAToken(s)
    wb.save('PythonZuExcelWithPressure.xlsx')

def CpPressureToXl101(data, atoken, s):
    wb = load_workbook(filename=r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\chargeTimeDataset.xlsx')
    FeuchteTbl = wb.worksheets[0]
    FeedbackTbl = wb.worksheets[1]
    CablePrs = searchXL(FeuchteTbl, "cablePressure")[1]
    j = 2
    for element in data:
        print(element)
        CPdata = PullPressure(element, atoken, s)
        searchStr = str(element['uniqueId'])
        Xcp = searchXL(FeuchteTbl, searchStr)[0]
        if Xcp != "notFound":
            if str(element['uniqueId'])[17] == "1":
                print("CP 1 found")
                i = 0
            elif str(element['uniqueId'])[17] == "2":
                print("CP 2 found")
                i = 1
            else:
                print("error CP ID besides 1 or 2 found")
                break
            print(CPdata)
            if CPdata['message'] != None:
                FeuchteTbl.cell(row=Xcp, column=CablePrs).value = float(CPdata['message']['idents'][16]['value'])
            else:
                FeuchteTbl.cell(row=Xcp, column=CablePrs).value = float(404)
            FeuchteTbl.cell(row=Xcp, column=CablePrs + 1).value = str(element['masterData']['chargePointName'])
            FeuchteTbl.cell(row=Xcp, column=CablePrs + 2).value = str(element['firmwareVersion'])
            FeedbackMessage = "Found in Row:" + str(Xcp)
        else:
            FeedbackMessage = "NotFound"
        FeedbackTbl.cell(row=j, column=1).value = str(element['uniqueId'])[:2]
        FeedbackTbl.cell(row=j, column=2).value = str(element['masterData']['chargingFacilities'][0]['power'])
        FeedbackTbl.cell(row=j, column=3).value = str(element['uniqueId'])
        FeedbackTbl.cell(row=j, column=4).value = str(element['masterData']['chargePointName'])
        FeedbackTbl.cell(row=j, column=5).value = str(element['firmwareVersion'])
        FeedbackTbl.cell(row=j, column=6).value = str(element['uniqueId'])[13:]
        FeedbackTbl.cell(row=j, column=7).value = FeedbackMessage
        j = j + 1
        if j % 10 == 0:
            atoken = GetAToken(s)
    wb.save(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\chargeTimeDataset.xlsx')

def GetAToken(s):
    url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC"
    refreshT(s)
    with open('token2.txt', 'r') as jsonf:
        data = json.load(jsonf)
        print("vergleich")
        print(data['refresh_token'])
    atoken = 'Bearer ' + data['access_token']
    return atoken

if __name__ == "__main__":
    i = 0
    s = requests.session()
    #url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC&status=ACTIVE&status=FAULTED&status=INACTIVE"
    atoken = GetAToken(s)
    #AllStandorteURL = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=5000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC"
    #AllCBXcp = BackReqTem(atoken, AllStandorteURL, s, 2, )
    #with open("UsableDestinationsDaily.txt", 'w') as f:
        #json.dump(AllCBXcp, f)

    f = open("UsableDestinationsDaily.txt", 'r')
    CPplaces = json.load(f)
    for CP in CPplaces:
        #print(CP)
        print("id:",CP['uniqueId'], "name:",CP['masterData']['chargePointName'])
    #for element in CPplaces:
        #Update(element, atoken, s)
    print("update fertig")
    atoken = GetAToken(s)
    #Cable pressure get written into a XL this is the Standart option
    #CpPressureToXl(CPplaces, atoken, s)

    #Cable pressure gets written in one Time needed excel that is only so special accaitions
    CpPressureToXl101(CPplaces, atoken, s)

    os.startfile('PythonZuExcelWithPressure.xlsx')
    # wb = load_workbook(filename='PythonZuExcel.xlsx')
    # FeuchteTbl = wb.worksheets[0]
    # LastCol = findTodayCol(FeuchteTbl)
    # for i in range(1,100):
    #     print(FeuchteTbl.cell(row=i, column=LastCol).value)