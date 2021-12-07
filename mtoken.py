import json
from openpyxl import Workbook,load_workbook

def showcontent():
    with open('Info.json', 'r') as c:
        obj = json.load(c)
        print(obj['content'][1]['masterData']['chargingFacilities'][0]['power'])
        #for elements in obj['content']:
        for stuff in obj['content'][1]['masterData']['chargingFacilities']:
            print(stuff['power'])

def showallcontent():
    with open('Info.json', 'r') as c:
        obj = json.load(c)
        print(obj['content'][1]['masterData']['chargingFacilities'][0]['power'])
        print(obj['content'][1])

        for elements in obj['content']:
            if int(elements['masterData']['chargingFacilities'][0]['power']) < 350:
                print( str(elements['masterData']['chargingFacilities'][0]['power']) + ":" + str(elements['uniqueId']) + ":" + str(elements['masterData']['chargePointName']))

def pythontoexcel():
    Exceldatei = Workbook()
    Tabellenblatt = Exceldatei.active
    Tabellenblatt['A1'] = 10
    Tabellenblatt.cell(row=2, column=2).value = 100
    Exceldatei.save("PythonZuExcel.xlsx")
def loadexcel():
    wb = load_workbook(filename='PythonZuExcel.xlsx')
    Tabellenblatt = wb.active
    Tabellenblatt['A3'] = 99
    wb.save('PythonZuExcel.xlsx')
def jsondirect():


if __name__ == '__main__':
    loadexcel()