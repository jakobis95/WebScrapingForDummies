import requests
import json
import time
from openpyxl import Workbook
from A2WorkingSkrips.CBXStatusUpdate import  BackendRequestTemplate as BackReqTem
from A2WorkingSkrips.CBXStatusUpdate import  authLoopRequest


def TransactionDownload(Filename, atoken, CPid, s, size=50):
    #url = "https://api.chargepoint-management.com/status/connectionStatusList"
    url = "https://api.chargepoint-management.com/chargepoint/v1/transaction/list/filter?page=0&size=" + str(size) + "&sort=startDate,desc"
    headers = {
        'authority': 'api.chargepoint-management.com',
        'method' :  'POST',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Authorization': atoken,
        'content-length' : '1594',
        'content-type': 'application/json',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36'
    }
    #cpoIds müssen mit denen aus dem Backend übereinstimmen, wenn 403 als fehler kommt dann einfach mal im Backend diese Payload kopieren und austauschen dann sollte es wieder gehen
    payload = {"cpoIds":[2,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,64,65,66,67,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208,209,211,212,213,214,215,217,218,219,220,221,222,223,224,225,226,227,228,229,230,231,232,233,234,235,236,237,238,239,240,241,242,243,244,245,246,247,248,249,250,251,252,253,254,255,256,257,258,259,260,261,262,263,264,265,266,267,268,269,270,271,272,273,274,275,276,277,278,279,280,281,282,283,284,285,286,287,288,289,290,291,292,293,294,295,296,297,298,299,301,302,303,304,305,306,307,308,309,310,311,312,313,314,315,316,317,318,319,320,321,322,323,324,325,326,327,328,329,330,331,332,333,334,335,336,337,338,339,340,341,342,343,344,345,346,347,348,349,350,351,352,353,354,355,356,357,358,359,360,361,362,363,364,365,366,367,368,369,370,371,373,374,375,376,378,379,380,381,382,383,384,385,387,388,389,390,391,392,393,394,395,396,397,398,399,400,401,402,403,404,405,406,407,408,409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425,426,427,428,429,430,431,432,433,434,435,436,437,438,439,440,441,442,443,444,445,446,447,448,449,450,451,452,453,454],"chargePointIds":[]}
    payload['chargePointIds'].append(CPid)

    p = s.post(url, headers=headers, verify=False, json=payload)
    #print(p.json())
    #print(p)
    # with open(Filename, 'w') as f:
    #     f.write(p.text)
    return p.json()

def Create_XLSX_From_Json(Filename, XLSXname):
    wb = Workbook()
    ws = wb.worksheets[0]
    r = 1
    col = 1
    with open(Filename, 'r') as f:
        obj = json.load(f)
        for cpTransactions in obj['content']:
            if r != 1:
                for index in cpTransactions:
                    print(cpTransactions[index])
                    ws.cell(row=r, column=col).value = str(cpTransactions[index])
                    col = col + 1
                col = 1
            else:
                for index in cpTransactions:
                    print("first row, header is created for column %d", col)
                    ws.cell(row=r, column=col).value = str(index)
                    col = col + 1
                col = 1
                r = r + 1
                for index in cpTransactions:
                    print(cpTransactions[index])
                    ws.cell(row=r, column=col).value = str(cpTransactions[index])
                    col = col + 1
                col = 1

            col = 1
            r = r + 1
    wb.save(XLSXname)
def Get_cpoIds():
    #Copy of the cpoIds payload from the Backend
    cpoIds = {2,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,64,65,66,67,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,158,159,160,161,162,163,164,165,166,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187,188,189,190,191,192,193,194,195,196,197,198,199,200,201,202,203,204,205,206,207,208,209,211,212,213,214,215,217,218,219,220,221,222,223,224,225,226,227,228,229,230,231,232,233,234,235,236,237,238,239,240,241,242,243,244,245,246,247,248,249,250,251,252,253,254,255,256,257,258,259,260,261,262,263,264,265,266,267,268,269,270,271,272,273,274,275,276,277,278,279,280,281,282,283,284,285,286,287,288,289,290,291,292,293,294,295,296,297,298,299,301,302,303,304,305,306,307,308,309,310,311,312,313,314,315,316,317,318,319,320,321,322,323,324,325,326,327,328,329,330,331,332,333,334,335,336,337,338,339,340,341,342,343,344,345,346,347,348,349,350,351,352,353,354,355,356,357,358,359,360,361,362,363,364,365,366,367,368,369,370,371,373,374,375,376,378,379,380,381,382,383,384,385,386,387,388,389,390,391,392,393,394,395,396,397,398,399,400,401,402,403,404,405,406,407,408,409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425}
    print(cpoIds)
    return cpoIds
def Update_CBXCP_list(CPsListname, atoken, s):
    AllStandorteURL = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=5000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC"
    AllCBXcp = BackReqTem(atoken, AllStandorteURL, s, 2,)
    #AllCBXcpJSON = AllCBXcp.json()  # alle CBX chargepoints in einer Tabelle mit verschiedenen anderen Metadaten
    with open(CPsListname.txt, 'w') as f:
        json.dump(AllCBXcp, f)

def Update_CBX_Transaction_DB(Session, Filename,CPsListname, atoken):
    i=0
    Gesamt = 0

    with open(Filename, 'r') as f:
        x = json.load(f)
    LastID = x['content'][-1]['chargePointId']
    print(LastID)

    with open(CPsListname, 'r') as f:
        obj = json.load(f)
    for CP in obj:
        i = i+1
    objectCount = i
    print("Es gibt ",objectCount ," CPs in der Liste")
    i=0
    #Findet die Lezte ChargePointId in Json datei um von dort weitere abzufragen
    for CP in obj:
        if CP['uniqueId'] == LastID:
            break
        else:
            i = i+1
    lastIDIndex = i
    #LastIDIndex wird um 1 erhöht um den Nächsten Chargepoint abzufragen
    firstNewIndex = lastIDIndex + 2
    i=0

    for index in range(firstNewIndex, objectCount):
        i = i+1
        CP = obj[index]
        payload = CP['uniqueId']
        if i < objectCount:
            TransactionDataCP = TransactionDownload(Filename, atoken, payload, Session, size=100)
            print("Standort ",i,"/",objectCount,":",CP['uniqueId'], " hat ",TransactionDataCP['totalElements'], " Transactions")
            Gesamt = Gesamt + TransactionDataCP['totalElements']
            print(Gesamt)
            if TransactionDataCP['totalElements'] >= 100:
                #time.sleep(5)
                TransactionDataCP = TransactionDownload(Filename, atoken, payload, Session, size=TransactionDataCP['totalElements'])
            if 'status' in TransactionDataCP:
                print("an error occoured with :", payload)
                time.sleep(5)
                TransactionDataCP = TransactionDownload(Filename, atoken, payload, Session, size=1000)
                if 'status' in TransactionDataCP:
                    print("an error occoured with :", payload)
                else:
                    JSONappend(Filename, TransactionDataCP)
            else:
                JSONappend(Filename, TransactionDataCP)
            atoken = refreshT(s)
        else: break
    print(i)
def JSONappend(Filename, JsonData):
    i=0
    # load Main Data File
    with open(Filename, 'r') as f:
        x = json.load(f)

    for Transaction in x['content']:
        i = i+1
    print("Transactions befor in x %d", i)
    i = 0
    for Transaction in JsonData['content']:
        x['content'].append(Transaction)

    for Transaction in x['content']:
        i = i+1
    print("Transactions in x %d", i)

    with open(Filename, 'w') as f:
        json.dump(x, f)

def create_empty_json(Session, atoken, Filename):

    pjson = TransactionDownload(Filename, atoken, "US9110000362_CPN01", Session, size=1000)
    print(pjson)
    with open(Filename, 'w') as f:
        json.dump(pjson, f)

    with open(Filename, 'r') as f:
        filedata = json.load(f)
    print(filedata)

def refreshT(s): # Hier wird der Token refreshed
    #die verwendete Json datei wird aus dem Browser kopiert
    with open('DataFiles/refreshtoken.txt', 'r') as jsonf:
        data = json.load(jsonf)
        #print(data['refresh_token'])


    url = 'https://login.chargepoint-management.com/auth/realms/PAG/protocol/openid-connect/token'
    headers = {
        'authority' : 'api.chargepoint-management.com',
        'accept' : '*/*',
        'accept-encoding' : 'gzip, deflate, br',
        'accept-language' : 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin' : 'https://www.chargepoint-management.com',
        'Referer':'https://www.chargepoint-management.com/',
        'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Cookie': 'AUTH_SESSION_ID=a4f12be7-2104-43c3-85a5-b62176eb82e3.e3b7f9c6-acc3-4abe-7b3a-4903; AUTH_SESSION_ID_LEGACY=a4f12be7-2104-43c3-85a5-b62176eb82e3.e3b7f9c6-acc3-4abe-7b3a-4903; KEYCLOAK_SESSION=PAG/12ca747b-237d-433c-afcc-9757ffe96d14/a4f12be7-2104-43c3-85a5-b62176eb82e3; KEYCLOAK_SESSION_LEGACY=PAG/12ca747b-237d-433c-afcc-9757ffe96d14/a4f12be7-2104-43c3-85a5-b62176eb82e3; KEYCLOAK_IDENTITY=eyJhbGciOiJIUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJiODJmYzRkZC1iOGI5LTQyODQtYjZlNy0wMTg3NTRlYWFiMzAifQ.eyJleHAiOjE2Mzg4NDgyNTEsImlhdCI6MTYzODgxMjI1MSwianRpIjoiZGFiYzMyYmQtYjJlOS00Njg2LWJkOTQtMGE3YWVmNDY0YmIyIiwiaXNzIjoiaHR0cHM6Ly9sb2dpbi5jaGFyZ2Vwb2ludC1tYW5hZ2VtZW50LmNvbS9hdXRoL3JlYWxtcy9QQUciLCJzdWIiOiIxMmNhNzQ3Yi0yMzdkLTQzM2MtYWZjYy05NzU3ZmZlOTZkMTQiLCJ0eXAiOiJTZXJpYWxpemVkLUlEIiwic2Vzc2lvbl9zdGF0ZSI6ImE0ZjEyYmU3LTIxMDQtNDNjMy04NWE1LWI2MjE3NmViODJlMyIsInN0YXRlX2NoZWNrZXIiOiJ4bWNfcHdCMVFobThOczNQSkFrNW91cDROeFlTUmN2UUhkWnh4LVFzWFFRIn0.fcif5iM_sOKPwJEoIQ3lYFsKvP0p_i8MASjq_q09FEA; KEYCLOAK_IDENTITY_LEGACY=eyJhbGciOiJIUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJiODJmYzRkZC1iOGI5LTQyODQtYjZlNy0wMTg3NTRlYWFiMzAifQ.eyJleHAiOjE2Mzg4NDgyNTEsImlhdCI6MTYzODgxMjI1MSwianRpIjoiZGFiYzMyYmQtYjJlOS00Njg2LWJkOTQtMGE3YWVmNDY0YmIyIiwiaXNzIjoiaHR0cHM6Ly9sb2dpbi5jaGFyZ2Vwb2ludC1tYW5hZ2VtZW50LmNvbS9hdXRoL3JlYWxtcy9QQUciLCJzdWIiOiIxMmNhNzQ3Yi0yMzdkLTQzM2MtYWZjYy05NzU3ZmZlOTZkMTQiLCJ0eXAiOiJTZXJpYWxpemVkLUlEIiwic2Vzc2lvbl9zdGF0ZSI6ImE0ZjEyYmU3LTIxMDQtNDNjMy04NWE1LWI2MjE3NmViODJlMyIsInN0YXRlX2NoZWNrZXIiOiJ4bWNfcHdCMVFobThOczNQSkFrNW91cDROeFlTUmN2UUhkWnh4LVFzWFFRIn0.fcif5iM_sOKPwJEoIQ3lYFsKvP0p_i8MASjq_q09FEA'
        }
    payload ={
        'grant_type':'refresh_token',
        'refresh_token' : data["refresh_token"],
        'client_id' : 'cpoc-frontend'
        }


    p = s.post(url, headers=headers, verify=False, data = payload )
    #print(p)

    with open('DataFiles/refreshtoken.txt', 'w') as f:
        f.write(p.text)

    with open('DataFiles/refreshtoken.txt', 'r') as jsonf:
        data = json.load(jsonf)
        #print("vergleich")
        #print( data['refresh_token'])

    return 'Bearer ' + data['access_token']



if __name__ == '__main__':
    Filename = 'DataFiles/TransactionData.txt'
    Filename2 = 'TransactionData5.txt'
    XLSXname = 'TransactionData5.xlsx'
    CPsListname = 'AllCPs.text'

    #Create Session for Backend
    s = requests.session()

    authLoopRequest(s)

    with open('DataFiles/refreshtoken.txt', 'r') as jsonf:
        data = json.load(jsonf)
        print("vergleich")
        print( data['refresh_token'])
    atoken = 'Bearer ' + data['access_token']
    input = "input"
    while input != "n" or "y":
        input = input("Möchten Sie mit einem Leeren Datenset Starten? y fuer Yes und n fuer No")
    if input != "y":
        print("Daten werden ab dem letzten Eintrag in der letzten Datei fortgeführt")
    else:
        create_empty_json(s, atoken, Filename2)
        #Downloads a list of all active Chargebox Charge Points
        Update_CBXCP_list(CPsListname, atoken, s)
        print("Leere Datei wird erstellt")


    #creates a complete Transaction dataset
    Update_CBX_Transaction_DB(s,Filename2, CPsListname, atoken)
    Create_XLSX_From_Json(Filename2, XLSXname)
    f = open(CPsListname)
    CPList = json.load(f)
    for standort in CPList:
        print(standort["manufacturerModelId"]["name"])

