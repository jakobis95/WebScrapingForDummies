import json
import os
from A2WorkingSkrips.CBXStatusUpdate import authLoopRequest
from openpyxl import load_workbook, styles
import requests
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL
import time


def Update(CP,atoken,s):
    elements = CP

    headers = {
        'authority': 'api.chargepoint-management.com',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Authorization': atoken
    }

    # update Ladestatus
    url = ("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:12] + "_LMS01/request?lmsGlobalId=00000000000100030" +str(elements['uniqueId'])[17] + "0a&force=true")
    # url = ("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:12] + "_LMS01/request?lmsGlobalId=de911000016700030"+ str(elements['uniqueId'])[17] +"0a&force=true")
    s.get(url, headers=headers, verify=False)
    time.sleep(2)
def PullPressure(CP, atoken, s):
    elements = CP

    headers = {
        'authority': 'api.chargepoint-management.com',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
        'Origin': 'https://www.chargepoint-management.com',
        'Referer': 'https://www.chargepoint-management.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
        'Authorization': atoken
    }

    # get mesurements loading control
    url = ("https://api.chargepoint-management.com/maintenance/v1/measurements/" + str(elements['uniqueId'])[:12] + "_LMS01?lmsGlobalId=00000000000100030" +str(elements['uniqueId'])[17] + "0a")
    print(url)
    CPdata = s.get(url, headers=headers, verify=False)
    CPdataJ = CPdata.json()
    return CPdataJ
def CpPressureToXl(data, atoken, s):
    wb = load_workbook(filename='../CablepressureAnalysis/PythonZuExcel.xlsx')
    FeuchteTbl = wb.worksheets[0]
    FeedbackTbl = wb.worksheets[1]
    CablePrs1 = searchXL(FeuchteTbl, "cablePressure1")[1]
    CablePrs2 = searchXL(FeuchteTbl, "cablePressure2")[1]
    CablePrs = (CablePrs1, CablePrs2)
    j = 2
    for element in data:
        print(element)
        CPdata = PullPressure(element, atoken, s)
        searchStr = str(element['uniqueId'])[:17] + "1"
        print("searchstr:", searchStr)
        Xcp = searchXL(FeuchteTbl, searchStr)[0]
        if Xcp != "notFound" and CPdata['message']['idents'][8]['value'] != "State B - car connected":
            if str(element['uniqueId'])[17] == "1":
                print("CP 1 found")
                i = 0
            elif str(element['uniqueId'])[17] == "2":
                print("CP 2 found")
                i = 1
            else:
                print("error CP ID besides 1 or 2 found")
                break
            print(CPdata)
            if CPdata['message'] != None:
                FeuchteTbl.cell(row=Xcp, column=CablePrs[i]).value = CPdata['message']['idents'][16]['value']
            else:
                FeuchteTbl.cell(row=Xcp, column=CablePrs[i]).value = "None"
            FeedbackMessage = "Found in Row:" + str(Xcp)

        else:
            if CPdata['message']['idents'][8]['value'] != "State B - car connected":
                FeedbackMessage = "NotFound"
            else:
                FeedbackMessage = "Car Connected Measurement invalid"
        FeedbackTbl.cell(row=j, column=1).value = str(element['uniqueId'])[:2]
        FeedbackTbl.cell(row=j, column=2).value = str(element['masterData']['chargingFacilities'][0]['power'])
        FeedbackTbl.cell(row=j, column=3).value = str(element['uniqueId'])
        FeedbackTbl.cell(row=j, column=4).value = str(element['masterData']['chargePointName'])
        FeedbackTbl.cell(row=j, column=5).value = str(element['firmwareVersion'])
        FeedbackTbl.cell(row=j, column=6).value = str(element['uniqueId'])[13:]
        FeedbackTbl.cell(row=j, column=7).value = FeedbackMessage
        j = j + 1
        if j % 10 == 0:
            atoken = GetAToken(s)
    wb.save('PythonZuExcelWithPressure.xlsx')
def CpPressureToXlRAW(data, atoken, s):
    #Fill Style definition

    my_yellow = styles.colors.Color(rgb='ffff00')
    my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
    no_fill = styles.PatternFill(fill_type=None)

    wb = load_workbook(filename=r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\Cable_Pressure_Temp_Raw.xlsx')
    FeuchteTbl = wb['ChargingData']
    FeedbackTbl = wb['Feedback PressureInput']
    CablePrs = searchXL(FeuchteTbl, "cablePressure")[1]
    j = 2
    for element in data:
        print(element)
        CPdata = PullPressure(element, atoken, s)
        searchStr = str(element['uniqueId'])
        Xcp = searchXL(FeuchteTbl, searchStr)[0]
        if Xcp != "notFound" and CPdata['message'] != None :
            if CPdata['message']['idents'][8]['value'] != "State B - car connected":
                if str(element['uniqueId'])[17] == "1":
                    print("CP 1 found")
                    i = 0
                elif str(element['uniqueId'])[17] == "2":
                    print("CP 2 found")
                    i = 1
                else:
                    print("error CP ID besides 1 or 2 found")
                    break
                print(CPdata)
                Cable_Pressure = float(CPdata['message']['idents'][16]['value'])
                Cable_Cooling_Outlet_Temp = float(CPdata['message']['idents'][25]['value'])
                Cable_Cooling_Inlet_Temp = float(CPdata['message']['idents'][24]['value'])
                Cable_Temperature = float(CPdata['message']['idents'][19]['value'])
                CCS_Temperature = float(CPdata['message']['idents'][12]['value'])
                Temperature_Cof = 0.013 # Temperaturkoeffizient/Grad des Druckes absolut auf 1,3Bar bei 10 Grad
                Mean_Cable_Temperature = (Cable_Temperature + Cable_Cooling_Outlet_Temp + Cable_Cooling_Inlet_Temp)/3
                Temperature_Corrected_Cable_Pressure = (20 - Mean_Cable_Temperature) * Temperature_Cof + Cable_Pressure
                print("Druck Differenz",(Mean_Cable_Temperature - 20) * Temperature_Cof)
                print("Temperaturdifferenz:", Mean_Cable_Temperature - 20)
                print("Temperaturbereinigter Kabeldruck",Temperature_Corrected_Cable_Pressure)
                if CPdata['message'] != None:
                    FeuchteTbl.cell(row=Xcp, column=CablePrs).value = Cable_Pressure
                    FeuchteTbl.cell(row=Xcp, column=3).value = Cable_Cooling_Inlet_Temp
                    FeuchteTbl.cell(row=Xcp, column=4).value = Cable_Cooling_Outlet_Temp
                    FeuchteTbl.cell(row=Xcp, column=5).value = Cable_Temperature
                    FeuchteTbl.cell(row=Xcp, column=6).value = CCS_Temperature
                else:
                    FeuchteTbl.cell(row=Xcp, column=CablePrs).value = float(404)
                FeuchteTbl.cell(row=Xcp, column=CablePrs + 1).value = str(element['masterData']['chargePointName'])
                FeuchteTbl.cell(row=Xcp, column=CablePrs + 2).value = str(element['firmwareVersion'])
                FeedbackMessage = "Found in Row:" + str(Xcp)
                #FeedbackTbl.cell(row=j, column=1).fill = my_fill
                FeuchteTbl.cell(row=Xcp, column=1).fill = my_fill
            else:
                FeedbackMessage = "Car Connected Measurement invalid"
        else:
                FeedbackMessage = "NotFound"

        FeedbackTbl.cell(row=j, column=1).value = str(element['uniqueId'])[:2]
        FeedbackTbl.cell(row=j, column=2).value = str(element['masterData']['chargingFacilities'][0]['power'])
        FeedbackTbl.cell(row=j, column=3).value = str(element['uniqueId'])
        FeedbackTbl.cell(row=j, column=4).value = str(element['masterData']['chargePointName'])
        FeedbackTbl.cell(row=j, column=5).value = str(element['firmwareVersion'])
        FeedbackTbl.cell(row=j, column=6).value = str(element['uniqueId'])[13:]
        FeedbackTbl.cell(row=j, column=7).value = FeedbackMessage
        j = j + 1
        if j % 10 == 0:
            atoken = GetAToken(s)
    wb.save(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\Cable_Pressure_Temp_Raw.xlsx')

def CpPressureToXl101(data, atoken, s):
    #Fill Style definition

    my_yellow = styles.colors.Color(rgb='ffff00')
    my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
    no_fill = styles.PatternFill(fill_type=None)

    wb = load_workbook(filename=r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V10.2.xlsx')
    FeuchteTbl = wb['ChargingData']
    FeedbackTbl = wb['Feedback PressureInput']
    CablePrs = searchXL(FeuchteTbl, "cablePressure")[1]
    j = 2
    for element in data:
        print(element)
        CPdata = PullPressure(element, atoken, s)
        searchStr = str(element['uniqueId'])
        Xcp = searchXL(FeuchteTbl, searchStr)[0]
        if Xcp != "notFound" and CPdata['message'] != None :
            if CPdata['message']['idents'][8]['value'] != "State B - car connected":
                if str(element['uniqueId'])[17] == "1":
                    print("CP 1 found")
                    i = 0
                elif str(element['uniqueId'])[17] == "2":
                    print("CP 2 found")
                    i = 1
                else:
                    print("error CP ID besides 1 or 2 found")
                    break
                print(CPdata)
                Cable_Pressure = float(CPdata['message']['idents'][16]['value'])
                Cable_Cooling_Outlet_Temp = float(CPdata['message']['idents'][25]['value'])
                Cable_Cooling_Inlet_Temp = float(CPdata['message']['idents'][24]['value'])
                Cable_Temperature = float(CPdata['message']['idents'][19]['value'])
                Temperature_Cof = 0.013 # Temperaturkoeffizient/Grad des Druckes absolut auf 1,3Bar bei 10 Grad
                Mean_Cable_Temperature = (Cable_Temperature + Cable_Cooling_Outlet_Temp + Cable_Cooling_Inlet_Temp)/3
                Temperature_Corrected_Cable_Pressure = (20 - Mean_Cable_Temperature) * Temperature_Cof + Cable_Pressure
                print("Druck Differenz",(Mean_Cable_Temperature - 20) * Temperature_Cof)
                print("Temperaturdifferenz:", Mean_Cable_Temperature - 20)
                print("Temperaturbereinigter Kabeldruck",Temperature_Corrected_Cable_Pressure)
                if CPdata['message'] != None:
                    FeuchteTbl.cell(row=Xcp, column=CablePrs).value = Temperature_Corrected_Cable_Pressure
                else:
                    FeuchteTbl.cell(row=Xcp, column=CablePrs).value = float(404)
                FeuchteTbl.cell(row=Xcp, column=CablePrs + 1).value = str(element['masterData']['chargePointName'])
                FeuchteTbl.cell(row=Xcp, column=CablePrs + 2).value = str(element['firmwareVersion'])
                FeedbackMessage = "Found in Row:" + str(Xcp)
                #FeedbackTbl.cell(row=j, column=1).fill = my_fill
                FeuchteTbl.cell(row=Xcp, column=1).fill = my_fill
            else:
                FeedbackMessage = "Car Connected Measurement invalid"
        else:
                FeedbackMessage = "NotFound"

        FeedbackTbl.cell(row=j, column=1).value = str(element['uniqueId'])[:2]
        FeedbackTbl.cell(row=j, column=2).value = str(element['masterData']['chargingFacilities'][0]['power'])
        FeedbackTbl.cell(row=j, column=3).value = str(element['uniqueId'])
        FeedbackTbl.cell(row=j, column=4).value = str(element['masterData']['chargePointName'])
        FeedbackTbl.cell(row=j, column=5).value = str(element['firmwareVersion'])
        FeedbackTbl.cell(row=j, column=6).value = str(element['uniqueId'])[13:]
        FeedbackTbl.cell(row=j, column=7).value = FeedbackMessage
        j = j + 1
        if j % 10 == 0:
            atoken = GetAToken(s)
    wb.save(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\220215_Seriennummern_LS_Ladekabel_V10.2.xlsx')

def GetAToken(s):
    url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC"
    authLoopRequest(s)
    with open('DataFiles/refreshtoken.txt', 'r') as jsonf:
        data = json.load(jsonf)
        print("vergleich")
        print(data['refresh_token'])
    atoken = 'Bearer ' + data['access_token']
    return atoken

if __name__ == "__main__":
    i = 0
    s = requests.session()
    #url = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=1000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC&status=ACTIVE&status=FAULTED&status=INACTIVE"
    atoken = GetAToken(s)
    #AllStandorteURL = "https://api.chargepoint-management.com/chargepoint/chargepoints/list?page=0&size=5000&sort=masterData.chargePointName,asc&masterData.chargingFacilities.powerType=DC"
    #AllCBXcp = BackReqTem(atoken, AllStandorteURL, s, 2, )
    #with open("UsableDestinationsDaily.txt", 'w') as f:
        #json.dump(AllCBXcp, f)

    f = open("DataFiles/UsableDestinationsDaily.txt", 'r')
    CPplaces = json.load(f)
    for CP in CPplaces:
        #print(CP)
        print("id:",CP['uniqueId'], "name:",CP['masterData']['chargePointName'])
    #for element in CPplaces:
        #Update(element, atoken, s)
    print("update fertig")
    atoken = GetAToken(s)
    #Cable pressure get written into a XL this is the Standart option
    #CpPressureToXl(CPplaces, atoken, s)

    #Cable pressure gets written in one Time needed excel that is only for special accaitions
    #CpPressureToXl101(CPplaces, atoken, s)
    CpPressureToXlRAW(CPplaces, atoken, s)

    os.startfile(r'C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\Cable_Pressure_Temp_Raw.xlsx')
    # wb = load_workbook(filename='PythonZuExcel.xlsx')
    # FeuchteTbl = wb.worksheets[0]
    # LastCol = findTodayCol(FeuchteTbl)
    # for i in range(1,100):
    #     print(FeuchteTbl.cell(row=i, column=LastCol).value)