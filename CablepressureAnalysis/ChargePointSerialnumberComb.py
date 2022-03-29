import json
from openpyxl import load_workbook
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL


if __name__ == "__main__":
    #variables
    xlsxPfad = r"C:\Users\FO4A5OY\OneDrive - Dr. Ing. h.c. F. Porsche AG\CablepressureAnalisys\SeriennummerStandortIDComb.xlsx"
    wb = load_workbook(filename=xlsxPfad)
    WB = wb["Tabelle1"]

    #Column serielnumber Chargepoints
    #columnSerialNr = searchXL(WB, 'Ladesäule')[1]
    columnSerialNr = 11
    data = open("../A2WorkingSkrips/DataFiles/UsableDestinationsDaily.txt")
    Udesti = json.load(data)
    CityM1 = "City"
    RowM1 = "notFound"
    SerialNumber = "Keine"
    for element in Udesti:
        if RowM1 != "notFound" and CityM1 == element['address']['city']:
            Row = RowM1 + 1
        else:
            Row = searchXL(WB, element['address']['city'])[0]
        if Row != "notFound":
            SerialNumber = WB.cell(column = columnSerialNr, row = Row ).value
            #print(element['serialNumber'], " City:", element['address']['city'], " seriennummer:",SerialNumber)
            RowM1 = Row
            CityM1 = element['address']['city']
        print(element['serialNumber'], " City:", element['address']['city'], " seriennummer:", SerialNumber)

        data = open("../A2WorkingSkrips/DataFiles/UsableDestinationsDaily.txt")
        Udesti = json.load(data)