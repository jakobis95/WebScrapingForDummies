from openpyxl import Workbook
import openpyxl


file = "Lennart.xlsx"
#wb = openpyxl.Workbook(file)
wb = openpyxl.load_workbook(filename=file)
ws = wb.worksheets[0]

RowNumber = ws.max_row
print("Es wurden " + str(RowNumber) + " Reihen gefunden" )
for i in range(RowNumber-1):
    TeilenummerRaw = ws.cell(row=i+1, column=1).value
    if len(TeilenummerRaw) > 19 :
        TeilenummerRaw = TeilenummerRaw[5:]
        TeilenummerRaw = str(TeilenummerRaw[0:3] + "." +TeilenummerRaw[3:6] + "."+TeilenummerRaw[6:9] + "."+ TeilenummerRaw[9:12] + "." + TeilenummerRaw[12:14] + "-" +TeilenummerRaw[14])
        print(TeilenummerRaw)
    else:
        TeilenummerRaw = TeilenummerRaw[5:]
        print("wurde nicht veränder :" + TeilenummerRaw)
    ws.cell(row=i+1, column=2).value = TeilenummerRaw
wb.save('Lennart.xlsx')