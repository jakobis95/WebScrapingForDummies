from openpyxl import load_workbook, styles
import os
import datetime
def takeRow(elem):
    return elem[1]

def returnPDFdata(xlsxPfad):

    wb = load_workbook(filename=xlsxPfad, keep_vba=True)
    IBN_S = wb.worksheets[0]
    maxRow = IBN_S.max_row
    maxCol = IBN_S.max_column
    No = 0
    dataRows = []
    cell = []
    for row in range(1, maxRow, 1):
        dataColumn = []
        for column in range(1, maxCol, 1):
            if IBN_S.cell(row=row, column=column).value:
                No = No + 1
                if isinstance(IBN_S.cell(row=row, column=column).value, datetime.datetime):
                    cell = IBN_S.cell(row=row, column=column).value
                    cell = cell.date()
                else:
                    cell = IBN_S.cell(row=row, column=column).value
                dataColumn.append(cell)
        dataRows.append(dataColumn)
    i = 0;
    for row in dataRows:
        if (i < 12 or i >= 44) and i < 97:
            print(i, ":", end= "")
            for column in row:
                print(column, end="; ")
            print("")
        i = i + 1
    return dataRows

if __name__ == "__main__":
    UserName = os.getlogin()
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Documents\\teesside cbx commissioning.xlsm"
    #xlsxPfad = "C:/Users/AJ2MSGR/OneDrivea - Dr.Ing.h.c.F.Porsche AG/IBN_Kastenfinden.xlsm"
    returnPDFdata(xlsxPfad)
