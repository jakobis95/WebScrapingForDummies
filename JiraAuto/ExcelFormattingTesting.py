from openpyxl import Workbook
import openpyxl
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from A3SupportingGeneralFunctions.NavigateInExcel import searchXL

if __name__ == "__main__":
    UserName = os.getlogin()
    xlsxPfad = "C:\\Users\\" + str(UserName) + "\\Downloads\\IBN_SANDbox3.xlsx"
    wb = load_workbook(filename=xlsxPfad)
    ws = wb["StatusKurz"]
    CW = datetime.today().isocalendar()
    TodayColumnString = "Full/Part KW" + str(CW[1])
    TodayCell = searchXL(ws, TodayColumnString)  # findet jetzt die heutige Spalte
    TodayColumn = TodayCell[1]
    TodayRow = TodayCell[0]
    FirstCellIndex = ws.cell(row=TodayRow, column=TodayColumn).coordinate
    LastCellIndex = ws.cell(row=ws.max_row, column=TodayColumn).coordinate

    green_text = Font(color="003300")
    green_fill = PatternFill(bgColor="99CC00")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="yes", dxf=dxf)
    formula = "NOT(ISERROR(SEARCH(\"yes\"," + str(FirstCellIndex) +")))"
    area = "",str(FirstCellIndex),":", str(LastCellIndex), ""
    rule.formula = [formula]
    ws.conditional_formatting.add('AV12:AV40', rule)

    # yellow_text = Font(color="FF6600")
    # yellow_fill = PatternFill(bgColor="FFCC00")
    # dxf = DifferentialStyle(font=yellow_text, fill=yellow_fill)
    # rule = Rule(type="containsText", operator="containsText", text="no", dxf=dxf)
    # formula = "NOT(ISERROR(SEARCH(\"yes\"," + str(FirstCellIndex) + ")))"
    # rule.formula = [formula]
    # ws.conditional_formatting.add('A1:F40', rule)

    wb.save(xlsxPfad)
    os.startfile(xlsxPfad)